from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from .models import (
    AnnualLoanSnapshot,
    AnnualPropertyCost,
    AnnualPropertySnapshot,
    ImportRun,
    Lease,
    LeasePerson,
    Loan,
    Property,
    RentPeriod,
    Tenant,
    Unit,
)
from .services import backfill_property_snapshots


def dec(value, default="0.00") -> Decimal:
    if value is None or value == "":
        return Decimal(default)
    if isinstance(value, str) and value.startswith("="):
        return Decimal(default)
    return Decimal(str(value)).quantize(Decimal("0.01"))


def share(value) -> Decimal:
    if value is None or value == "":
        return Decimal("1.0")
    return Decimal(str(value)).quantize(Decimal("0.000001"))


def clean_text(value) -> str:
    return str(value or "").strip()


@transaction.atomic
def import_master_immos(path_or_file, year: int) -> ImportRun:
    wb_values = load_workbook(path_or_file, data_only=True)
    warnings: list[str] = []
    row_mappings: dict[str, list[dict]] = {"properties": [], "tenants": []}

    imported_properties: dict[str, Property] = {}
    rent_ws = wb_values["Rentabilität"]
    for row in range(7, rent_ws.max_row + 1):
        address = clean_text(rent_ws.cell(row, 1).value)
        if not address or address.startswith("Summe"):
            continue
        owner_share = share(rent_ws.cell(row, 3).value)
        purchase_price = dec(rent_ws.cell(row, 8).value)
        debt = dec(rent_ws.cell(row, 10).value)
        total_down_payment = max(Decimal("0.00"), purchase_price - debt).quantize(Decimal("0.01"))
        property_obj, _ = Property.objects.update_or_create(
            address=address,
            defaults={
                "name": address.split(",")[0],
                "ownership_share": owner_share,
                "purchase_price": purchase_price,
                "cash_invested_at_purchase": total_down_payment,
                "recurring_expense_amount": dec(rent_ws.cell(row, 14).value),
            },
        )
        imported_properties[address] = property_obj
        value = dec(rent_ws.cell(row, 8).value)
        snapshot, _ = AnnualPropertySnapshot.objects.update_or_create(
            property=property_obj,
            year=year,
            defaults={"property_value": value, "valuation_source": "Imported purchase price"},
        )
        AnnualPropertyCost.objects.update_or_create(
            snapshot=snapshot,
            category=AnnualPropertyCost.OTHER,
            defaults={"amount": Decimal("0.00"), "notes": "Running costs are stored on the property as yearly recurring expense"},
        )
        rate = Decimal(str(rent_ws.cell(row, 12).value or 0)).quantize(Decimal("0.000001"))
        if debt:
            interest = (debt * rate).quantize(Decimal("0.01"))
            loan, _ = Loan.objects.update_or_create(
                property=property_obj,
                name=f"{property_obj.name} loan",
                defaults={
                    "original_amount": debt,
                    "default_interest_rate": rate,
                    "default_monthly_payment": (interest / Decimal("12")).quantize(Decimal("0.01")),
                    "start_date": date(year, 1, 1),
                },
            )
            AnnualLoanSnapshot.objects.update_or_create(
                loan=loan,
                year=year,
                defaults={
                    "opening_balance": debt,
                    "closing_balance": debt,
                    "interest_paid": interest,
                    "principal_paid": Decimal("0.00"),
                    "interest_rate": rate,
                    "monthly_payment": (interest / Decimal("12")).quantize(Decimal("0.01")),
                    "debt_service": interest,
                },
            )
        row_mappings["properties"].append({"row": row, "address": address, "property_id": property_obj.pk})

    for property_obj in imported_properties.values():
        backfill_property_snapshots(property_obj, through_year=year)

    raw_ws = wb_values["Rohdaten"]
    for row in range(3, raw_ws.max_row + 1):
        last_name = clean_text(raw_ws.cell(row, 1).value)
        if last_name.lower() == "archiv":
            break
        if not last_name:
            continue
        first_name = clean_text(raw_ws.cell(row, 2).value)
        address = clean_text(raw_ws.cell(row, 3).value)
        property_obj = imported_properties.get(address)
        if not property_obj:
            property_obj, _ = Property.objects.get_or_create(address=address, defaults={"name": address.split(",")[0] or "Imported property"})
            warnings.append(f"Created property from Rohdaten row {row}: {address}")

        tenant, _ = Tenant.objects.update_or_create(
            last_name=last_name,
            first_name=first_name,
            defaults={
                "email": clean_text(raw_ws.cell(row, 5).value),
                "phone": clean_text(raw_ws.cell(row, 4).value),
                "notes": clean_text(raw_ws.cell(row, 6).value),
            },
        )
        floor = clean_text(raw_ws.cell(row, 10).value)
        unit_label = f"Floor {floor}" if floor else f"Unit {row - 2}"
        unit, _ = Unit.objects.get_or_create(property=property_obj, label=unit_label, defaults={"floor": floor})
        lease, _ = Lease.objects.update_or_create(
            unit=unit,
            tenant=tenant,
            defaults={"start_date": date(year, 1, 1), "is_active": True},
        )
        LeasePerson.objects.get_or_create(
            lease=lease,
            person=tenant,
            role=LeasePerson.PRIMARY,
            defaults={
                "move_in_date": lease.start_date,
                "move_out_date": lease.end_date,
                "is_contract_signer": True,
            },
        )
        cold = dec(raw_ws.cell(row, 7).value)
        utilities = dec(raw_ws.cell(row, 9).value)
        total = dec(raw_ws.cell(row, 11).value, default=str(cold + utilities))
        rent_period = RentPeriod.objects.filter(lease=lease, effective_start=date(year, 1, 1)).first()
        if rent_period:
            rent_period.cold_rent = cold
            rent_period.utility_prepayment = utilities
            rent_period.total_rent = total or cold + utilities
            rent_period.save()
        else:
            RentPeriod.objects.create(
                lease=lease,
                effective_start=date(year, 1, 1),
                cold_rent=cold,
                utility_prepayment=utilities,
                total_rent=total or cold + utilities,
                notes="Imported initial rent period",
            )
        row_mappings["tenants"].append({"row": row, "tenant_id": tenant.pk, "property_id": property_obj.pk})

    source_name = getattr(path_or_file, "name", None) or Path(str(path_or_file)).name
    return ImportRun.objects.create(source_name=source_name, warnings=warnings, row_mappings=row_mappings)
