from __future__ import annotations

import shutil
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.db import connection
from django.db.models import Sum
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import AnnualLoanSnapshot, AnnualPropertySnapshot, Loan, Property, ReportExport
from .services import current_debt_from_annual_snapshot, money, unit_rent_overview


ZERO = Decimal("0.00")


@dataclass(frozen=True)
class CurrentRentUnitRow:
    property: Property
    unit_label: str
    unit_floor: str
    tenant_name: str
    lease_start: date | None
    monthly_cold_rent: Decimal
    monthly_utility_prepayment: Decimal
    monthly_total_rent: Decimal
    annual_cold_rent: Decimal
    annual_utility_prepayment: Decimal
    annual_total_rent: Decimal
    owner_annual_total_rent: Decimal


@dataclass(frozen=True)
class CurrentRentPropertyRow:
    property: Property
    monthly_cold_rent: Decimal
    monthly_utility_prepayment: Decimal
    monthly_total_rent: Decimal
    annual_cold_rent: Decimal
    annual_utility_prepayment: Decimal
    annual_total_rent: Decimal
    owner_annual_total_rent: Decimal


@dataclass(frozen=True)
class CurrentRentOverview:
    as_of: date
    properties: list[CurrentRentPropertyRow]
    units: list[CurrentRentUnitRow]
    total_monthly_cold_rent: Decimal
    total_monthly_utility_prepayment: Decimal
    total_monthly_total_rent: Decimal
    total_annual_cold_rent: Decimal
    total_annual_utility_prepayment: Decimal
    total_annual_total_rent: Decimal
    total_owner_annual_total_rent: Decimal


@dataclass(frozen=True)
class BankRealEstateRow:
    property: Property
    object_type: str
    address: str
    ownership_share: Decimal
    construction_year: int | None
    living_area_sqm: Decimal | None
    total_area_sqm: Decimal | None
    property_value: Decimal
    annual_cold_rent: Decimal
    uses_purchase_price_fallback: bool
    uses_total_area_fallback: bool
    has_no_area_fallback: bool


@dataclass(frozen=True)
class BankLoanRow:
    loan: Loan
    property: Property
    initial_loan_value: Decimal
    lender: str
    current_loan_amount: Decimal
    fixed_until_year: int | None
    monthly_payment: Decimal
    interest_rate: Decimal


@dataclass(frozen=True)
class BankFinancingOverview:
    as_of: date
    real_estate: list[BankRealEstateRow]
    loans: list[BankLoanRow]


@dataclass(frozen=True)
class BankFinancingPreviewRow:
    property: Property
    property_value: Decimal
    annual_cold_rent: Decimal
    debt: Decimal
    monthly_payment: Decimal
    interest_rate: Decimal | None
    fixed_until_year: int | None


@dataclass(frozen=True)
class BankFinancingPreview:
    overview: BankFinancingOverview
    rows: list[BankFinancingPreviewRow]
    warnings: list[str]


def _record_export(export_type: str, title: str, file_name: str, **extra) -> None:
    defaults = {"title": title, **extra}
    ReportExport.objects.update_or_create(export_type=export_type, file_name=file_name, defaults=defaults)


def current_rent_overview(as_of: date | None = None) -> CurrentRentOverview:
    as_of = as_of or date.today()
    unit_rows = []
    property_rows = []
    total_monthly_cold = ZERO
    total_monthly_utility = ZERO
    total_monthly_total = ZERO
    total_annual_cold = ZERO
    total_annual_utility = ZERO
    total_annual_total = ZERO
    total_owner_annual_total = ZERO

    for property_obj in Property.objects.prefetch_related("units", "units__leases__tenant", "units__leases__rent_periods").all():
        property_monthly_cold = ZERO
        property_monthly_utility = ZERO
        property_monthly_total = ZERO
        property_annual_cold = ZERO
        property_annual_utility = ZERO
        property_annual_total = ZERO
        property_owner_annual_total = ZERO

        for unit in property_obj.units.all():
            overview = unit_rent_overview(unit, as_of)
            lease = overview.active_lease
            annual_cold = money(overview.cold_rent * 12)
            annual_utility = money(overview.utility_prepayment * 12)
            annual_total = money(overview.total_rent * 12)
            owner_annual_total = money(annual_total * property_obj.ownership_share)

            unit_rows.append(
                CurrentRentUnitRow(
                    property=property_obj,
                    unit_label=unit.label,
                    unit_floor=unit.floor,
                    tenant_name=str(lease.tenant) if lease else "Vacant",
                    lease_start=lease.start_date if lease else None,
                    monthly_cold_rent=overview.cold_rent,
                    monthly_utility_prepayment=overview.utility_prepayment,
                    monthly_total_rent=overview.total_rent,
                    annual_cold_rent=annual_cold,
                    annual_utility_prepayment=annual_utility,
                    annual_total_rent=annual_total,
                    owner_annual_total_rent=owner_annual_total,
                )
            )

            property_monthly_cold += overview.cold_rent
            property_monthly_utility += overview.utility_prepayment
            property_monthly_total += overview.total_rent
            property_annual_cold += annual_cold
            property_annual_utility += annual_utility
            property_annual_total += annual_total
            property_owner_annual_total += owner_annual_total

        property_row = CurrentRentPropertyRow(
            property=property_obj,
            monthly_cold_rent=money(property_monthly_cold),
            monthly_utility_prepayment=money(property_monthly_utility),
            monthly_total_rent=money(property_monthly_total),
            annual_cold_rent=money(property_annual_cold),
            annual_utility_prepayment=money(property_annual_utility),
            annual_total_rent=money(property_annual_total),
            owner_annual_total_rent=money(property_owner_annual_total),
        )
        property_rows.append(property_row)

        total_monthly_cold += property_row.monthly_cold_rent
        total_monthly_utility += property_row.monthly_utility_prepayment
        total_monthly_total += property_row.monthly_total_rent
        total_annual_cold += property_row.annual_cold_rent
        total_annual_utility += property_row.annual_utility_prepayment
        total_annual_total += property_row.annual_total_rent
        total_owner_annual_total += property_row.owner_annual_total_rent

    return CurrentRentOverview(
        as_of=as_of,
        properties=property_rows,
        units=unit_rows,
        total_monthly_cold_rent=money(total_monthly_cold),
        total_monthly_utility_prepayment=money(total_monthly_utility),
        total_monthly_total_rent=money(total_monthly_total),
        total_annual_cold_rent=money(total_annual_cold),
        total_annual_utility_prepayment=money(total_annual_utility),
        total_annual_total_rent=money(total_annual_total),
        total_owner_annual_total_rent=money(total_owner_annual_total),
    )


def bank_financing_overview(as_of: date | None = None) -> BankFinancingOverview:
    as_of = as_of or date.today()
    rent_overview = current_rent_overview(as_of)
    rent_by_property_id = {row.property.pk: row for row in rent_overview.properties}
    properties = (
        Property.objects.select_related("administration")
        .prefetch_related("units")
        .annotate(unit_area_sum=Sum("units__area_sqm"))
        .order_by("name")
    )
    real_estate_rows = []
    for property_obj in properties:
        administration = getattr(property_obj, "administration", None)
        total_area = administration.total_building_area_sqm if administration else None
        unit_area_sum = property_obj.unit_area_sum
        living_area = unit_area_sum if unit_area_sum is not None and unit_area_sum > ZERO else total_area
        latest_value = (
            AnnualPropertySnapshot.objects.filter(property=property_obj)
            .order_by("-year")
            .values_list("property_value", flat=True)
            .first()
        )
        rent_row = rent_by_property_id.get(property_obj.pk)
        uses_total_area_fallback = (unit_area_sum is None or unit_area_sum <= ZERO) and total_area is not None
        real_estate_rows.append(
            BankRealEstateRow(
                property=property_obj,
                object_type=property_obj.get_object_type_display() if property_obj.object_type else "",
                address=property_obj.display_address,
                ownership_share=property_obj.ownership_share,
                construction_year=administration.construction_year if administration else None,
                living_area_sqm=living_area,
                total_area_sqm=total_area,
                property_value=money(latest_value if latest_value is not None else property_obj.purchase_price),
                annual_cold_rent=rent_row.annual_cold_rent if rent_row else ZERO,
                uses_purchase_price_fallback=latest_value is None,
                uses_total_area_fallback=uses_total_area_fallback,
                has_no_area_fallback=(unit_area_sum is None or unit_area_sum <= ZERO) and total_area is None,
            )
        )

    loan_rows = []
    for loan in Loan.objects.select_related("property").prefetch_related("annual_snapshots").order_by("property__name", "name"):
        snapshot = _bank_relevant_loan_snapshot(loan, as_of)
        if snapshot:
            current_amount = _bank_current_loan_amount(snapshot, as_of)
            monthly_payment = snapshot.monthly_payment if snapshot.monthly_payment else loan.default_monthly_payment
            interest_rate = snapshot.interest_rate if snapshot.interest_rate else loan.default_interest_rate
            fixed_until = snapshot.rate_reset_date.year if snapshot.rate_reset_date else None
        else:
            current_amount = loan.original_amount
            monthly_payment = loan.default_monthly_payment
            interest_rate = loan.default_interest_rate
            fixed_until = None
        loan_rows.append(
            BankLoanRow(
                loan=loan,
                property=loan.property,
                initial_loan_value=money(loan.original_amount),
                lender=loan.lender,
                current_loan_amount=money(current_amount),
                fixed_until_year=fixed_until,
                monthly_payment=money(monthly_payment),
                interest_rate=interest_rate or ZERO,
            )
        )

    return BankFinancingOverview(as_of=as_of, real_estate=real_estate_rows, loans=loan_rows)


def bank_financing_preview(as_of: date | None = None) -> BankFinancingPreview:
    overview = bank_financing_overview(as_of)
    loans_by_property_id: dict[int, list[BankLoanRow]] = {}
    for loan in overview.loans:
        loans_by_property_id.setdefault(loan.property.pk, []).append(loan)

    rows = []
    warnings = []
    for real_estate in overview.real_estate:
        property_loans = loans_by_property_id.get(real_estate.property.pk, [])
        debt = money(sum((loan.current_loan_amount for loan in property_loans), ZERO))
        monthly_payment = money(sum((loan.monthly_payment for loan in property_loans), ZERO))
        weighted_interest_total = sum((loan.current_loan_amount * loan.interest_rate for loan in property_loans), ZERO)
        interest_rate = weighted_interest_total / debt if debt else None
        fixed_years = [loan.fixed_until_year for loan in property_loans if loan.fixed_until_year]

        rows.append(
            BankFinancingPreviewRow(
                property=real_estate.property,
                property_value=real_estate.property_value,
                annual_cold_rent=real_estate.annual_cold_rent,
                debt=debt,
                monthly_payment=monthly_payment,
                interest_rate=interest_rate,
                fixed_until_year=min(fixed_years) if fixed_years else None,
            )
        )

        if real_estate.uses_purchase_price_fallback:
            warnings.append(f"Property value missing for {real_estate.property.name}. Purchase price will be used.")
        if real_estate.uses_total_area_fallback:
            warnings.append(f"Living area missing for {real_estate.property.name}. Total property area will be used instead.")
        if real_estate.has_no_area_fallback:
            warnings.append(f"Living area missing for {real_estate.property.name}. No fallback area is available.")

    return BankFinancingPreview(overview=overview, rows=rows, warnings=warnings)


def export_bank_financing_workbook(as_of: date | None = None) -> bytes:
    overview = bank_financing_overview(as_of)
    wb = Workbook()

    real_estate = wb.active
    real_estate.title = "Real Estate"
    real_estate.append(["As of", overview.as_of.isoformat()])
    real_estate.append([])
    real_estate.append([
        "Object name",
        "Object type",
        "Address",
        "Ownership share",
        "Year built",
        "Living area sqm",
        "Total property area sqm",
        "Property value",
        "Annual cold rent",
    ])
    for row in overview.real_estate:
        real_estate.append([
            row.property.name,
            row.object_type,
            row.address,
            float(row.ownership_share),
            row.construction_year,
            float(row.living_area_sqm) if row.living_area_sqm is not None else None,
            float(row.total_area_sqm) if row.total_area_sqm is not None else None,
            float(row.property_value),
            float(row.annual_cold_rent),
        ])

    loans = wb.create_sheet("Loans")
    loans.append(["As of", overview.as_of.isoformat()])
    loans.append([])
    loans.append([
        "Property",
        "Loan name",
        "Initial loan value",
        "Lender",
        "Current loan amount",
        "Interest fixed until year",
        "Total monthly payment",
        "Interest rate",
    ])
    for row in overview.loans:
        loans.append([
            row.property.name,
            row.loan.name,
            float(row.initial_loan_value),
            row.lender,
            float(row.current_loan_amount),
            row.fixed_until_year,
            float(row.monthly_payment),
            float(row.interest_rate),
        ])

    _format_bank_workbook(wb)
    stream = BytesIO()
    wb.save(stream)
    file_name = f"bank-financing-overview-{overview.as_of.isoformat()}.xlsx"
    _record_export("bank_financing_excel", "Bank Financing Overview Excel", file_name)
    return stream.getvalue()


def bank_financing_pdf(as_of: date | None = None) -> bytes:
    overview = bank_financing_overview(as_of)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), title="Bank Financing Overview", leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Bank Financing Overview", styles["Title"]),
        Paragraph(f"As of {overview.as_of:%d.%m.%Y}", styles["Normal"]),
        Spacer(1, 12),
    ]

    real_estate_rows = [["Object", "Type", "Address", "Share", "Built", "Living area", "Total property area", "Value", "Cold rent p.a."]]
    for row in overview.real_estate:
        real_estate_rows.append([
            row.property.name,
            row.object_type or "-",
            row.address or "-",
            f"{row.ownership_share:.2%}",
            row.construction_year or "-",
            _sqm(row.living_area_sqm),
            _sqm(row.total_area_sqm),
            _eur(row.property_value),
            _eur(row.annual_cold_rent),
        ])
    elements.append(Paragraph("Real Estate", styles["Heading2"]))
    elements.append(_styled_table(real_estate_rows, font_size=7))
    elements.append(Spacer(1, 14))

    loan_rows = [["Property", "Loan", "Initial value", "Lender", "Current amount", "Fixed until", "Monthly payment", "Interest"]]
    for row in overview.loans:
        loan_rows.append([
            row.property.name,
            row.loan.name,
            _eur(row.initial_loan_value),
            row.lender or "-",
            _eur(row.current_loan_amount),
            row.fixed_until_year or "",
            _eur(row.monthly_payment),
            f"{row.interest_rate:.2%}",
        ])
    elements.append(Paragraph("Loans", styles["Heading2"]))
    elements.append(_styled_table(loan_rows, font_size=7))

    doc.build(elements)
    file_name = f"bank-financing-overview-{overview.as_of.isoformat()}.pdf"
    _record_export("bank_financing_pdf", "Bank Financing Overview PDF", file_name)
    return buffer.getvalue()


def create_database_backup() -> str:
    settings.APP_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    source = connection.settings_dict["NAME"]
    target = settings.APP_BACKUP_DIR / f"app-{datetime.now():%Y%m%d-%H%M%S}.db"
    shutil.copy2(source, target)
    _record_export("backup", "Database backup", target.name)
    return str(target)


def restore_database_backup(uploaded_file) -> str:
    settings.APP_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    staged = settings.APP_BACKUP_DIR / f"restore-upload-{datetime.now():%Y%m%d-%H%M%S}.db"
    with staged.open("wb") as target:
        for chunk in uploaded_file.chunks():
            target.write(chunk)
    with staged.open("rb") as source:
        if source.read(16) != b"SQLite format 3\x00":
            staged.unlink(missing_ok=True)
            raise ValueError("The uploaded file is not a SQLite database.")
    connection.close()
    with sqlite3.connect(staged) as src, sqlite3.connect(connection.settings_dict["NAME"]) as dst:
        src.backup(dst)
    _record_export("backup", "Database restored", staged.name)
    return str(staged)


def _bank_relevant_loan_snapshot(loan: Loan, as_of: date) -> AnnualLoanSnapshot | None:
    snapshots = sorted(loan.annual_snapshots.all(), key=lambda snapshot: snapshot.year, reverse=True)
    for snapshot in snapshots:
        if snapshot.year <= as_of.year:
            return snapshot
    return None


def _bank_current_loan_amount(snapshot: AnnualLoanSnapshot, as_of: date) -> Decimal:
    if snapshot.year == as_of.year:
        return current_debt_from_annual_snapshot(snapshot.opening_balance, snapshot.closing_balance, snapshot.year, as_of)
    return money(snapshot.closing_balance)


def _format_bank_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for cell in ws[3]:
            cell.font = cell.font.copy(bold=True)
        for column_cells in ws.columns:
            values = [str(cell.value) for cell in column_cells if cell.value not in (None, "")]
            width = min(max([len(value) for value in values] or [10]) + 2, 42)
            ws.column_dimensions[column_cells[0].column_letter].width = width
    real_estate = wb["Real Estate"]
    for row in real_estate.iter_rows(min_row=4, min_col=4, max_col=9):
        row[0].number_format = "0.00%"
        for cell in row[4:]:
            cell.number_format = '#,##0.00'
    loans = wb["Loans"]
    for row in loans.iter_rows(min_row=4, min_col=3, max_col=8):
        row[0].number_format = '#,##0.00'
        row[2].number_format = '#,##0.00'
        row[4].number_format = '#,##0.00'
        row[5].number_format = "0.00%"


def _styled_table(rows, font_size: int = 9):
    table = Table(rows, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16324f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def _pct_or_na(value):
    if value is None:
        return "n/a"
    return f"{value:.2%}"


def _eur(value: Decimal) -> str:
    return f"€ {value:,.2f}"


def _sqm(value: Decimal | None) -> str:
    if value is None:
        return "-"
    return f"{value:,.2f} sqm"
