import base64
import importlib
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook

from .exports import bank_financing_overview, bank_financing_preview, export_bank_financing_workbook, current_rent_overview
from .forms import (
    LeasePeopleForm,
    LoanBalanceTableForm,
    PotentialDealCreateForm,
    PotentialDealForm,
    PotentialDealOptimizerForm,
    PotentialFinancingScenarioForm,
    PropertyCreateForm,
    PropertyHistoryTableForm,
    RentChangeForm,
    TenantChangeForm,
    UnitAdministrationForm,
    UnitWithTenantForm,
)
from .models import (
    AnnualLoanSnapshot,
    AnnualPropertyCost,
    AnnualPropertySnapshot,
    AnnualPortfolioTax,
    AppSettings,
    Contact,
    LandlordProfile,
    Lease,
    LeasePerson,
    Loan,
    PotentialDeal,
    PotentialFinancingScenario,
    Property,
    PropertyAdministration,
    ReportExport,
    RentPeriod,
    Tenant,
    Unit,
    UnitAdministration,
    UnitContact,
    UnitLandRegistry,
    UnitTechnicalInfo,
)
from .services import annual_rent_totals, backfill_property_snapshots, current_debt_from_annual_snapshot, dashboard_data_quality, loan_performance_rows, optimize_potential_deal_scenario, portfolio_performance, potential_deal_metrics, potential_deal_portfolio_comparison, potential_deal_scenario_comparisons, property_performance, ratio


SIGNATURE_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgG"
    "aWjR9awAAAABJRU5ErkJggg=="
)


class FinanceCalculationTests(TestCase):
    def setUp(self):
        self.property = Property.objects.create(
            name="Test House",
            address="Test Street 1",
            ownership_share=Decimal("0.5"),
            purchase_price=Decimal("200000.00"),
            cash_invested_at_purchase=Decimal("20000.00"),
        )
        self.unit = Unit.objects.create(property=self.property, label="Unit 1")
        self.tenant = Tenant.objects.create(first_name="Ada", last_name="Lovelace")
        self.lease = Lease.objects.create(unit=self.unit, tenant=self.tenant, start_date=date(2026, 1, 1))

    def test_rent_prorates_mid_year_changes(self):
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), effective_end=date(2026, 6, 30), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 7, 1), cold_rent=Decimal("1100"), utility_prepayment=Decimal("220"), total_rent=Decimal("1320"))

        totals = annual_rent_totals(self.property, 2026)

        self.assertEqual(totals.cold_rent, Decimal("12600.00"))
        self.assertEqual(totals.utility_prepayment, Decimal("2520.00"))
        self.assertEqual(totals.total_rent, Decimal("15120.00"))

    def test_overlapping_rent_periods_are_rejected(self):
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), effective_end=date(2026, 6, 30), cold_rent=Decimal("1000"))
        with self.assertRaises(ValidationError):
            RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 6, 1), effective_end=date(2026, 12, 31), cold_rent=Decimal("1200"))

    def test_property_performance_calculates_core_kpis(self):
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        snapshot = AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000"), vacancy_loss=Decimal("500"))
        AnnualPropertyCost.objects.create(snapshot=snapshot, category=AnnualPropertyCost.MAINTENANCE, amount=Decimal("1500"))
        loan = Loan.objects.create(property=self.property, name="Main loan", original_amount=Decimal("200000"))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2026, opening_balance=Decimal("200000"), closing_balance=Decimal("196000"), interest_paid=Decimal("4000"), principal_paid=Decimal("4000"))

        metrics = property_performance(snapshot)

        self.assertEqual(metrics.property_value, Decimal("125000.00"))
        self.assertEqual(metrics.annual_total_rent, Decimal("7200.00"))
        self.assertEqual(metrics.recurring_expense, Decimal("750.00"))
        self.assertEqual(metrics.operating_costs, Decimal("750.00"))
        self.assertEqual(metrics.noi, Decimal("4750.00"))
        self.assertEqual(metrics.debt_service, Decimal("4000.00"))
        self.assertEqual(metrics.cashflow, Decimal("2750.00"))
        self.assertEqual(metrics.free_cashflow, Decimal("750.00"))
        self.assertEqual(metrics.ltv, Decimal("0.7840"))
        self.assertEqual(metrics.purchase_cash_out, Decimal("10000.00"))
        self.assertEqual(metrics.annual_cash_in, Decimal("6000.00"))
        self.assertEqual(metrics.annual_cash_out, Decimal("3250.00"))
        self.assertEqual(metrics.annual_equity_build, Decimal("2000.00"))
        self.assertEqual(metrics.annual_owner_roi, Decimal("0.2750"))
        self.assertEqual(metrics.cumulative_cash_out, Decimal("13250.00"))
        self.assertEqual(metrics.cumulative_cashflow, Decimal("2750.00"))
        self.assertEqual(metrics.cumulative_free_cashflow, Decimal("750.00"))
        self.assertEqual(metrics.cumulative_owner_roi, Decimal("0.2750"))

    def test_property_performance_rows_stay_before_portfolio_tax(self):
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        snapshot = AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000"), vacancy_loss=Decimal("500"))
        AnnualPropertyCost.objects.create(snapshot=snapshot, category=AnnualPropertyCost.MAINTENANCE, amount=Decimal("1500"))
        loan = Loan.objects.create(property=self.property, name="Main loan", original_amount=Decimal("200000"))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2026, opening_balance=Decimal("200000"), closing_balance=Decimal("196000"), interest_paid=Decimal("4000"), principal_paid=Decimal("4000"))
        tax_settings = AppSettings.objects.create(effective_tax_rate=Decimal("0.250000"), tax_loss_benefit_enabled=True)

        before_tax = property_performance(snapshot, tax_mode="before", tax_settings=tax_settings)
        after_tax = property_performance(snapshot, tax_mode="after", tax_settings=tax_settings)

        self.assertEqual(before_tax.cashflow, Decimal("2750.00"))
        self.assertEqual(before_tax.free_cashflow, Decimal("750.00"))
        self.assertEqual(before_tax.annual_owner_roi, Decimal("0.2750"))
        self.assertEqual(after_tax.cashflow_before_tax, Decimal("2750.00"))
        self.assertEqual(after_tax.estimated_taxable_result, Decimal("2750.00"))
        self.assertEqual(after_tax.estimated_tax, Decimal("0.00"))
        self.assertEqual(after_tax.cashflow_after_tax, Decimal("2750.00"))
        self.assertEqual(after_tax.cashflow, Decimal("2750.00"))
        self.assertEqual(after_tax.free_cashflow, Decimal("750.00"))
        self.assertEqual(after_tax.annual_owner_roi, Decimal("0.2750"))
        self.assertEqual(after_tax.noi, before_tax.noi)
        self.assertEqual(after_tax.ltv, before_tax.ltv)
        self.assertEqual(after_tax.debt_service_coverage, before_tax.debt_service_coverage)
        self.assertEqual(after_tax.annual_equity_build, before_tax.annual_equity_build)

    def test_portfolio_performance_after_tax_applies_loss_benefit(self):
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        snapshot = AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000"))
        AnnualPropertyCost.objects.create(snapshot=snapshot, category=AnnualPropertyCost.MAINTENANCE, amount=Decimal("20000"))
        tax_settings = AppSettings.objects.create(effective_tax_rate=Decimal("0.250000"), tax_loss_benefit_enabled=True)

        metrics = portfolio_performance(2026, tax_mode="after", tax_settings=tax_settings)

        self.assertEqual(metrics.total_cashflow_before_tax, Decimal("-4000.00"))
        self.assertEqual(metrics.total_estimated_tax, Decimal("-1000.00"))
        self.assertEqual(metrics.total_cashflow_after_tax, Decimal("-3000.00"))
        self.assertEqual(metrics.total_cashflow, Decimal("-3000.00"))

    def test_portfolio_performance_after_tax_uses_portfolio_tax_bridge(self):
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        snapshot = AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000"))
        AnnualPropertyCost.objects.create(snapshot=snapshot, category=AnnualPropertyCost.MAINTENANCE, amount=Decimal("1000"))
        AnnualPortfolioTax.objects.create(year=2026, tax_deductible_costs=Decimal("1000.00"))
        tax_settings = AppSettings.objects.create(effective_tax_rate=Decimal("0.100000"), tax_loss_benefit_enabled=True)

        portfolio = portfolio_performance(2026, tax_mode="after", tax_settings=tax_settings)

        row = portfolio.rows[0]
        self.assertEqual(portfolio.total_cashflow_before_tax, row.cashflow_before_tax)
        self.assertEqual(portfolio.annual_tax_deductible_costs, Decimal("1000.00"))
        self.assertEqual(portfolio.portfolio_taxable_result, Decimal("4500.00"))
        self.assertEqual(portfolio.total_estimated_tax, Decimal("450.00"))
        self.assertEqual(portfolio.total_cashflow_after_tax, Decimal("5050.00"))
        self.assertEqual(portfolio.total_cashflow, Decimal("5050.00"))
        self.assertEqual(portfolio.total_free_cashflow, Decimal("5050.00"))
        self.assertEqual(portfolio.annual_owner_roi, ratio(portfolio.total_cashflow_after_tax, portfolio.total_purchase_cash_out))

    def test_property_recurring_expense_override_replaces_default(self):
        self.property.recurring_expense_amount = Decimal("2000.00")
        self.property.save()
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        snapshot = AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000"), vacancy_loss=Decimal("0"))

        metrics = property_performance(snapshot)

        self.assertEqual(metrics.recurring_expense, Decimal("1000.00"))
        self.assertEqual(metrics.operating_costs, Decimal("1000.00"))
        self.assertEqual(metrics.noi, Decimal("5000.00"))

    def test_migrated_zero_cost_keeps_recurring_expense_fallback(self):
        self.property.recurring_expense_amount = Decimal("2000.00")
        self.property.save()
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        snapshot = AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000"), vacancy_loss=Decimal("0"))
        AnnualPropertyCost.objects.create(snapshot=snapshot, category=AnnualPropertyCost.OTHER, amount=Decimal("0.00"), notes="Migrated to property yearly recurring expense")

        metrics = property_performance(snapshot)

        self.assertEqual(metrics.operating_costs, Decimal("1000.00"))

    def test_explicit_zero_yearly_cost_is_used_as_zero(self):
        self.property.recurring_expense_amount = Decimal("2000.00")
        self.property.save()
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        snapshot = AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000"), vacancy_loss=Decimal("0"))
        AnnualPropertyCost.objects.create(snapshot=snapshot, category=AnnualPropertyCost.OTHER, amount=Decimal("0.00"), notes="Yearly non-recoverable costs")

        metrics = property_performance(snapshot)

        self.assertEqual(metrics.operating_costs, Decimal("0.00"))
        self.assertEqual(metrics.noi, Decimal("6000.00"))

    def test_property_history_table_creates_missing_years_and_cost_rows(self):
        form = PropertyHistoryTableForm(
            data={
                "name": "Test House",
                "address": "Test Street 1",
                "ownership_share": "0.500000",
                "purchase_price": "200000.00",
                "cash_invested_at_purchase": "20000.00",
                "acquisition_date": "01.02.2024",
                "selected_year": "2026",
                "notes": "Updated",
                "property_value_2024": "200000.00",
                "non_recoverable_costs_2024": "1000.00",
                "vacancy_loss_2024": "0.00",
                "manual_rent_adjustment_2024": "0.00",
                "notes_2024": "",
                "property_value_2025": "205000.00",
                "non_recoverable_costs_2025": "1100.00",
                "vacancy_loss_2025": "100.00",
                "manual_rent_adjustment_2025": "50.00",
                "notes_2025": "Partial repair",
                "property_value_2026": "210000.00",
                "non_recoverable_costs_2026": "1200.00",
                "vacancy_loss_2026": "0.00",
                "manual_rent_adjustment_2026": "0.00",
                "notes_2026": "",
            },
            property_obj=self.property,
            selected_year=2026,
        )

        self.assertTrue(form.is_valid(), form.errors)
        property_obj = form.save()

        self.assertEqual(property_obj.acquisition_date, date(2024, 2, 1))
        self.assertEqual(list(property_obj.annual_snapshots.order_by("year").values_list("year", flat=True)), [2024, 2025, 2026])
        snapshot = property_obj.annual_snapshots.get(year=2025)
        self.assertEqual(snapshot.property_value, Decimal("205000.00"))
        self.assertEqual(snapshot.manual_rent_adjustment, Decimal("50.00"))
        cost = snapshot.costs.get(category=AnnualPropertyCost.OTHER)
        self.assertEqual(cost.amount, Decimal("1100.00"))
        self.assertEqual(cost.notes, "Yearly non-recoverable costs")

    def test_property_history_missing_values_default_to_purchase_price(self):
        self.property.acquisition_date = date(2020, 1, 1)
        self.property.save(update_fields=["acquisition_date"])
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000.00"))
        form = PropertyHistoryTableForm(property_obj=self.property, selected_year=2026)

        self.assertEqual(form.fields["property_value_2020"].initial, Decimal("200000.00"))
        self.assertEqual(form.fields["property_value_2025"].initial, Decimal("200000.00"))
        self.assertEqual(form.fields["property_value_2026"].initial, Decimal("250000.00"))

    def test_backfill_property_snapshots_uses_purchase_price_without_overwriting_manual_values(self):
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000.00"), valuation_source="Manual")
        loan = Loan.objects.create(property=self.property, name="Historical loan", original_amount=Decimal("100000.00"), start_date=date(2020, 1, 1))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2020, opening_balance=Decimal("100000.00"), closing_balance=Decimal("95000.00"))

        created = backfill_property_snapshots(self.property, through_year=2026)

        self.assertEqual(created, 6)
        values = dict(self.property.annual_snapshots.order_by("year").values_list("year", "property_value"))
        self.assertEqual(values[2020], Decimal("200000.00"))
        self.assertEqual(values[2025], Decimal("200000.00"))
        self.assertEqual(values[2026], Decimal("250000.00"))
        self.assertEqual(self.property.annual_snapshots.get(year=2020).valuation_source, "Default purchase price")

    def test_dashboard_data_quality_reports_complete_and_missing_inputs(self):
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000.00"), utility_prepayment=Decimal("200.00"), total_rent=Decimal("1200.00"))
        snapshot = AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000.00"))
        AnnualPropertyCost.objects.create(snapshot=snapshot, category=AnnualPropertyCost.OTHER, amount=Decimal("1000.00"), notes="Yearly non-recoverable costs")
        loan = Loan.objects.create(property=self.property, name="Main loan", original_amount=Decimal("100000.00"))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2026, opening_balance=Decimal("100000.00"), closing_balance=Decimal("95000.00"))
        broken_property = Property.objects.create(name="Broken House", ownership_share=Decimal("0.000000"), purchase_price=Decimal("0.00"))
        broken_unit = Unit.objects.create(property=broken_property, label="Unit A")
        broken_tenant = Tenant.objects.create(first_name="Missing", last_name="Rent")
        Lease.objects.create(unit=broken_unit, tenant=broken_tenant, start_date=date(2026, 1, 1))
        Loan.objects.create(property=broken_property, name="Missing balance", original_amount=Decimal("50000.00"))

        complete_quality = dashboard_data_quality(2026, [self.property.pk])
        complete_messages = [item.message for item in complete_quality.items]
        all_quality = dashboard_data_quality(2026)
        all_messages = [item.message for item in all_quality.items]

        self.assertIn("Rent history complete", complete_messages)
        self.assertIn("Loan balance entered", complete_messages)
        self.assertFalse(any(item.severity == "warning" for item in complete_quality.items))
        self.assertIn("Rent history missing for Broken House / Unit A.", all_messages)
        self.assertIn("Loan balance missing for Broken House / Missing balance.", all_messages)
        self.assertIn("Property value missing for Broken House. No purchase price fallback is available.", all_messages)
        self.assertIn("Non-recoverable costs missing for Broken House. Using 5% rent estimate.", all_messages)
        self.assertIn("Ownership share missing for Broken House. Your-share KPIs may be distorted.", all_messages)

    def test_property_create_form_saves_simple_and_advanced_fields(self):
        form = PropertyCreateForm(
            data={
                "name": "Quick House",
                "street_address": "Quick Street 1",
                "postal_code": "34117",
                "city": "Kassel",
                "object_type": Property.MULTI_FAMILY_HOUSE,
                "ownership_share": "0.500000",
                "purchase_price": "300000.00",
                "cash_invested_at_purchase": "60000.00",
                "acquisition_date": "15.01.2026",
                "recurring_expense_amount": "1200.00",
                "construction_year": "1998",
                "total_building_area_sqm": "220.50",
                "notes": "Simple-first setup",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        property_obj = form.save()
        administration = property_obj.administration

        self.assertEqual(property_obj.object_type, Property.MULTI_FAMILY_HOUSE)
        self.assertEqual(property_obj.address, "Quick Street 1\n34117 Kassel")
        self.assertEqual(property_obj.cash_invested_at_purchase, Decimal("60000.00"))
        self.assertEqual(property_obj.acquisition_date, date(2026, 1, 15))
        self.assertEqual(property_obj.recurring_expense_amount, Decimal("1200.00"))
        self.assertEqual(administration.construction_year, 1998)
        self.assertEqual(administration.total_building_area_sqm, Decimal("220.50"))

    def test_structured_address_migration_parser_handles_one_line_address(self):
        migration = importlib.import_module("portfolio.migrations.0026_landlord_profiles_structured_addresses")

        self.assertEqual(migration.split_address("Heestweg 2, 22143 Hamburg"), ("Heestweg 2", "22143", "Hamburg"))
        self.assertEqual(migration.split_address("Test Street 1\n34117 Kassel"), ("Test Street 1", "34117", "Kassel"))

    def test_unit_with_tenant_form_creates_unit_lease_tenant_and_rent(self):
        form = UnitWithTenantForm(
            data={
                "property": self.property.pk,
                "label": "Unit 2",
                "floor": "2",
                "area_sqm": "65.00",
                "notes": "Corner apartment",
                "tenant_mode": "new",
                "existing_tenant": "",
                "first_name": "Grace",
                "last_name": "Hopper",
                "email": "grace@example.com",
                "phone": "",
                "tenant_notes": "Prefers email",
                "lease_start_date": "01.03.2026",
                "cold_rent": "900.00",
                "utility_prepayment": "180.00",
                "total_rent": "",
                "heating_type": "Gas",
                "boiler_installation_info": "Installed 2022",
                "cellar_number": "C-2",
                "local_court": "Hamburg",
                "land_register_district": "District A",
                "sheet_number": "123",
                "plot_numbers": "45/6",
                "management_contact_name": "Care Office",
                "management_contact_email": "care@example.com",
                "management_contact_phone": "+49 40 123",
            },
            property_obj=self.property,
        )

        self.assertTrue(form.is_valid(), form.errors)
        unit = form.save()
        lease = unit.leases.get()
        rent = lease.rent_periods.get()

        self.assertEqual(unit.label, "Unit 2")
        self.assertEqual(lease.tenant.last_name, "Hopper")
        self.assertTrue(lease.is_active)
        lease_person = lease.people.get()
        self.assertEqual(lease_person.person.last_name, "Hopper")
        self.assertEqual(lease_person.role, LeasePerson.PRIMARY)
        self.assertTrue(lease_person.is_contract_signer)
        self.assertEqual(rent.effective_start, date(2026, 3, 1))
        self.assertEqual(rent.total_rent, Decimal("1080.00"))
        self.assertEqual(unit.notes, "Corner apartment")
        self.assertEqual(unit.administration.cellar_number, "C-2")
        self.assertEqual(unit.technical_info.heating_type, "Gas")
        self.assertEqual(unit.technical_info.boiler_installation_info, "Installed 2022")
        self.assertEqual(unit.land_registry.land_register_district, "District A")
        management_link = unit.contacts.select_related("contact").get()
        self.assertEqual(management_link.role, UnitContact.PROPERTY_MANAGEMENT)
        self.assertEqual(management_link.contact.name, "Care Office")
        self.assertEqual(management_link.contact.email, "care@example.com")

    def test_tenant_change_closes_old_lease_and_open_rent_period(self):
        old_rent = RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        old_person = LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=date(2026, 1, 1))
        form = TenantChangeForm(
            data={
                "tenant_mode": "new",
                "existing_tenant": "",
                "first_name": "Grace",
                "last_name": "Hopper",
                "email": "",
                "phone": "",
                "tenant_notes": "",
                "lease_start_date": "01.07.2026",
                "cold_rent": "1100.00",
                "utility_prepayment": "220.00",
                "total_rent": "",
            },
            unit=self.unit,
        )

        self.assertTrue(form.is_valid(), form.errors)
        new_lease = form.save()
        self.lease.refresh_from_db()
        old_rent.refresh_from_db()

        self.assertEqual(self.lease.end_date, date(2026, 6, 30))
        self.assertFalse(self.lease.is_active)
        old_person.refresh_from_db()
        self.assertEqual(old_person.move_out_date, date(2026, 6, 30))
        self.assertEqual(old_rent.effective_end, date(2026, 6, 30))
        self.assertEqual(new_lease.start_date, date(2026, 7, 1))
        self.assertTrue(new_lease.is_active)
        self.assertEqual(new_lease.people.get().role, LeasePerson.PRIMARY)
        self.assertEqual(new_lease.people.get().person.last_name, "Hopper")
        self.assertEqual(new_lease.rent_periods.get().total_rent, Decimal("1320.00"))

    def test_lease_people_form_adds_contract_tenant_and_child_without_changing_rent(self):
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        primary = LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=date(2026, 1, 1))
        co_tenant_form = LeasePeopleForm(
            data={
                "add_person_mode": "new",
                "existing_person": "",
                "first_name": "Grace",
                "last_name": "Hopper",
                "email": "grace@example.com",
                "phone": "",
                "birthday": "",
                "relationship_notes": "",
                "new_role": LeasePerson.CO_TENANT,
                "new_move_in_date": "01.01.2026",
                "new_move_out_date": "",
                "new_notes": "Second contract tenant",
                f"link_{primary.pk}_first_name": "Ada",
                f"link_{primary.pk}_last_name": "Lovelace",
                f"link_{primary.pk}_email": "",
                f"link_{primary.pk}_phone": "",
                f"link_{primary.pk}_birthday": "",
                f"link_{primary.pk}_relationship_notes": "",
                f"link_{primary.pk}_role": LeasePerson.PRIMARY,
                f"link_{primary.pk}_move_in_date": "01.01.2026",
                f"link_{primary.pk}_move_out_date": "",
                f"link_{primary.pk}_notes": "",
            },
            lease=self.lease,
        )

        self.assertTrue(co_tenant_form.is_valid(), co_tenant_form.errors)
        co_tenant_form.save()

        child_form = LeasePeopleForm(
            data={
                "add_person_mode": "new",
                "existing_person": "",
                "first_name": "Mini",
                "last_name": "Hopper",
                "email": "",
                "phone": "",
                "birthday": "01.05.2020",
                "relationship_notes": "Child of Grace",
                "new_role": LeasePerson.CHILD,
                "new_move_in_date": "01.01.2026",
                "new_move_out_date": "",
                "new_notes": "Lives in the unit",
                f"link_{primary.pk}_first_name": "Ada",
                f"link_{primary.pk}_last_name": "Lovelace",
                f"link_{primary.pk}_email": "",
                f"link_{primary.pk}_phone": "",
                f"link_{primary.pk}_birthday": "",
                f"link_{primary.pk}_relationship_notes": "",
                f"link_{primary.pk}_role": LeasePerson.PRIMARY,
                f"link_{primary.pk}_move_in_date": "01.01.2026",
                f"link_{primary.pk}_move_out_date": "",
                f"link_{primary.pk}_notes": "",
                f"link_{self.lease.people.get(person__last_name='Hopper', role=LeasePerson.CO_TENANT).pk}_first_name": "Grace",
                f"link_{self.lease.people.get(person__last_name='Hopper', role=LeasePerson.CO_TENANT).pk}_last_name": "Hopper",
                f"link_{self.lease.people.get(person__last_name='Hopper', role=LeasePerson.CO_TENANT).pk}_email": "grace@example.com",
                f"link_{self.lease.people.get(person__last_name='Hopper', role=LeasePerson.CO_TENANT).pk}_phone": "",
                f"link_{self.lease.people.get(person__last_name='Hopper', role=LeasePerson.CO_TENANT).pk}_birthday": "",
                f"link_{self.lease.people.get(person__last_name='Hopper', role=LeasePerson.CO_TENANT).pk}_relationship_notes": "",
                f"link_{self.lease.people.get(person__last_name='Hopper', role=LeasePerson.CO_TENANT).pk}_role": LeasePerson.CO_TENANT,
                f"link_{self.lease.people.get(person__last_name='Hopper', role=LeasePerson.CO_TENANT).pk}_move_in_date": "01.01.2026",
                f"link_{self.lease.people.get(person__last_name='Hopper', role=LeasePerson.CO_TENANT).pk}_move_out_date": "",
                f"link_{self.lease.people.get(person__last_name='Hopper', role=LeasePerson.CO_TENANT).pk}_notes": "Second contract tenant",
            },
            lease=self.lease,
        )

        self.assertTrue(child_form.is_valid(), child_form.errors)
        child_form.save()

        self.assertEqual(self.lease.people.filter(role__in=LeasePerson.CONTRACT_ROLES).count(), 2)
        child = self.lease.people.get(role=LeasePerson.CHILD)
        self.assertFalse(child.is_contract_signer)
        self.assertEqual(child.person.relationship_notes, "Child of Grace")
        self.assertEqual(annual_rent_totals(self.property, 2026).total_rent, Decimal("14400.00"))

    def test_lease_people_form_rejects_removing_last_contract_tenant(self):
        primary = LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=date(2026, 1, 1))
        form = LeasePeopleForm(
            data={
                "add_person_mode": "none",
                f"link_{primary.pk}_first_name": "Ada",
                f"link_{primary.pk}_last_name": "Lovelace",
                f"link_{primary.pk}_email": "",
                f"link_{primary.pk}_phone": "",
                f"link_{primary.pk}_birthday": "",
                f"link_{primary.pk}_relationship_notes": "",
                f"link_{primary.pk}_role": LeasePerson.PRIMARY,
                f"link_{primary.pk}_move_in_date": "01.01.2026",
                f"link_{primary.pk}_move_out_date": "",
                f"link_{primary.pk}_notes": "",
                f"link_{primary.pk}_remove": "on",
            },
            lease=self.lease,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("at least one contract tenant", str(form.errors))

    def test_rent_change_closes_previous_open_period(self):
        old_rent = RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        form = RentChangeForm(
            data={
                "effective_start": "01.07.2026",
                "cold_rent": "1100.00",
                "utility_prepayment": "220.00",
                "total_rent": "",
                "notes": "",
            },
            lease=self.lease,
        )

        self.assertTrue(form.is_valid(), form.errors)
        new_rent = form.save()
        old_rent.refresh_from_db()

        self.assertEqual(old_rent.effective_end, date(2026, 6, 30))
        self.assertEqual(new_rent.effective_start, date(2026, 7, 1))
        self.assertEqual(new_rent.total_rent, Decimal("1320.00"))

    def test_editing_rent_period_start_updates_lease_start_and_rent_calculation(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        rent = RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        person_link = LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=date(2026, 1, 1))

        response = self.client.post(
            reverse("rent_period_update", kwargs={"pk": rent.pk}),
            {
                "lease": self.lease.pk,
                "effective_start": "01.01.2020",
                "effective_end": "",
                "cold_rent": "1000.00",
                "utility_prepayment": "200.00",
                "total_rent": "1200.00",
                "notes": "",
            },
        )

        self.assertRedirects(response, reverse("unit_detail", kwargs={"pk": self.unit.pk}))
        self.lease.refresh_from_db()
        rent.refresh_from_db()
        self.assertEqual(rent.effective_start, date(2020, 1, 1))
        self.assertEqual(self.lease.start_date, date(2020, 1, 1))
        person_link.refresh_from_db()
        self.assertEqual(person_link.move_in_date, date(2020, 1, 1))
        self.assertEqual(annual_rent_totals(self.property, 2020).total_rent, Decimal("14400.00"))

        detail = self.client.get(f"{reverse('unit_detail', kwargs={'pk': self.unit.pk})}#history")
        self.assertContains(detail, "01.01.2020 - open")

    def test_property_detail_shows_unit_overview(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        self.lease.start_date = date.today()
        self.lease.save()
        RentPeriod.objects.create(lease=self.lease, effective_start=date.today(), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))

        response = self.client.get(reverse("property_detail", kwargs={"pk": self.property.pk}))

        self.assertContains(response, "Unit 1")
        self.assertContains(response, "Lovelace, Ada")
        self.assertContains(response, "Property Info")
        self.assertContains(response, "No photo")
        self.assertNotContains(response, "€1,000.00")
        self.assertNotContains(response, "€200.00")

    def test_simplified_navigation_moves_maintenance_into_more(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        response = self.client.get(reverse("dashboard"))

        self.assertContains(response, "Dashboard")
        self.assertContains(response, "Properties")
        self.assertContains(response, "Tenants")
        self.assertContains(response, "Loans")
        self.assertContains(response, "Deals")
        self.assertContains(response, "<summary>")
        self.assertContains(response, "More")
        self.assertContains(response, "Reference")
        self.assertContains(response, reverse("reference"))
        self.assertContains(response, reverse("potential_deal_list"))
        self.assertContains(response, reverse("exports"))
        self.assertContains(response, reverse("import_workbook"))
        self.assertContains(response, reverse("backup"))
        self.assertContains(response, reverse("settings"))

    def test_reference_page_requires_login(self):
        response = self.client.get(reverse("reference"))

        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_reference_page_renders_formula_document(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        response = self.client.get(reverse("reference"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "CasaFlow Reference")
        self.assertContains(response, "Cashflow = cold rent - operating costs - vacancy/loss - interest")
        self.assertContains(response, "Free Cashflow = Cashflow - principal repayment")
        self.assertContains(response, "Total Value Added = Free Cashflow + Equity Build + unrealized value gain")
        self.assertContains(response, "Workflow Mind Map")
        self.assertContains(response, "CasaFlow Modes")
        self.assertContains(response, "Portfolio Tracking")
        self.assertContains(response, "Portfolio Analysis")
        self.assertContains(response, "Portfolio Decisions")
        self.assertContains(response, "Existing properties")
        self.assertContains(response, "Financing scenarios")
        self.assertNotContains(response, "Portfolio trend over time")
        self.assertNotContains(response, "Reference guide")
        self.assertTrue((Path(__file__).resolve().parent.parent / "docs" / "casaflow_reference.md").exists())

    def test_potential_deal_form_accepts_full_property_inputs(self):
        form = PotentialDealForm(
            data={
                "name": "New House",
                "address": "Deal Street 1",
                "status": PotentialDeal.REVIEW,
                "purchase_price": "300000.00",
                "ownership_share": "0.500000",
                "expected_monthly_cold_rent": "1500.00",
                "expected_monthly_utility_prepayment": "300.00",
                "yearly_non_recoverable_costs": "",
                "notes": "Interesting location",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        deal = form.save()

        self.assertEqual(deal.purchase_price, Decimal("300000.00"))
        self.assertEqual(deal.ownership_share, Decimal("0.500000"))

    def test_potential_deal_create_form_creates_default_scenario(self):
        form = PotentialDealCreateForm(
            data={
                "name": "Simple Deal",
                "purchase_price": "300000.00",
                "expected_monthly_cold_rent": "1500.00",
                "owner_cash_out": "50000.00",
                "loan_amount": "250000.00",
                "interest_rate": "4.2500",
                "monthly_payment": "1250.00",
                "expected_monthly_utility_prepayment": "200.00",
                "yearly_non_recoverable_costs": "1800.00",
                "ownership_share": "0.500000",
                "buying_costs": "25000.00",
                "minimum_dscr": "1.25",
                "maximum_ltv": "0.800000",
                "notes": "Quick start deal",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        deal = form.save()
        scenario = deal.scenarios.get()

        self.assertEqual(deal.purchase_price, Decimal("300000.00"))
        self.assertEqual(deal.buying_costs, Decimal("25000.00"))
        self.assertEqual(deal.minimum_dscr, Decimal("1.25"))
        self.assertEqual(deal.maximum_ltv, Decimal("0.800000"))
        self.assertEqual(scenario.name, "Initial scenario")
        self.assertTrue(scenario.is_default)
        self.assertEqual(scenario.owner_cash_out, Decimal("50000.00"))
        self.assertEqual(scenario.loan_amount, Decimal("250000.00"))
        self.assertEqual(scenario.interest_rate, Decimal("0.042500"))
        self.assertEqual(scenario.monthly_payment, Decimal("1250.00"))

    def test_potential_financing_scenario_form_uses_percentage_input(self):
        deal = PotentialDeal.objects.create(name="New House", purchase_price=Decimal("300000.00"))
        form = PotentialFinancingScenarioForm(
            data={
                "name": "Commerzbank",
                "owner_cash_out": "40000,00",
                "loan_amount": "200000.00",
                "interest_rate": "5,0000",
                "monthly_payment": "9500",
                "maturity_notes": "",
                "notes": "",
            },
            deal=deal,
        )

        self.assertTrue(form.is_valid(), form.errors)
        scenario = form.save()

        self.assertEqual(scenario.interest_rate, Decimal("0.050000"))
        self.assertEqual(scenario.monthly_payment, Decimal("9500"))
        self.assertEqual(scenario.owner_cash_out, Decimal("40000.00"))
        self.assertEqual(form.fields["interest_rate"].prepare_value(scenario.interest_rate), "5.0000")
        self.assertEqual(form.fields["monthly_payment"].widget.input_type, "text")
        self.assertEqual(form.fields["owner_cash_out"].label, "Your Cash Invested")

    def test_potential_deal_metrics_use_default_costs_and_owner_share(self):
        deal = PotentialDeal.objects.create(
            name="New House",
            purchase_price=Decimal("300000.00"),
            ownership_share=Decimal("0.500000"),
            expected_monthly_cold_rent=Decimal("1500.00"),
            expected_monthly_utility_prepayment=Decimal("300.00"),
        )
        scenario = PotentialFinancingScenario.objects.create(deal=deal, name="Bank A", owner_cash_out=Decimal("40000.00"), loan_amount=Decimal("200000.00"), interest_rate=Decimal("0.030000"), monthly_payment=Decimal("1000.00"))

        metrics = potential_deal_metrics(deal, scenario)

        self.assertEqual(metrics.operating_costs, Decimal("900.00"))
        self.assertEqual(metrics.owner_operating_costs, Decimal("450.00"))
        self.assertEqual(metrics.owner_cash_out, Decimal("40000.00"))
        self.assertEqual(metrics.debt_service, Decimal("6000.00"))
        self.assertEqual(metrics.interest_cost, Decimal("3000.00"))
        self.assertEqual(metrics.principal_repayment_estimate, Decimal("3000.00"))
        self.assertEqual(metrics.annual_cash_in, Decimal("9000.00"))
        self.assertEqual(metrics.annual_cash_out, Decimal("3450.00"))
        self.assertEqual(metrics.cashflow, Decimal("5550.00"))
        self.assertEqual(metrics.free_cashflow, Decimal("2550.00"))
        self.assertEqual(metrics.annual_equity_build, Decimal("3000.00"))
        self.assertEqual(metrics.liquidity_cashflow, Decimal("2550.00"))
        self.assertEqual(metrics.monthly_cashflow, Decimal("462.50"))
        self.assertEqual(metrics.monthly_liquidity_cashflow, Decimal("212.50"))
        self.assertEqual(metrics.monthly_debt_service, Decimal("500.00"))
        self.assertEqual(metrics.monthly_principal_repayment_estimate, Decimal("250.00"))
        self.assertEqual(metrics.annual_owner_roi, Decimal("0.1388"))
        self.assertEqual(metrics.cash_on_cash_return, Decimal("0.0638"))
        self.assertEqual(metrics.principal_repayment_return, Decimal("0.0750"))
        self.assertEqual(metrics.gross_yield, Decimal("0.0600"))
        self.assertEqual(metrics.net_yield, Decimal("0.0570"))
        self.assertEqual(metrics.ltv, Decimal("0.6667"))
        self.assertEqual(metrics.dscr, Decimal("1.4250"))
        self.assertEqual(metrics.years_cold_rent_to_price, Decimal("16.6667"))

    def test_potential_deal_metrics_match_dashboard_cashflow_structure(self):
        deal = PotentialDeal.objects.create(
            name="New House",
            purchase_price=Decimal("300000.00"),
            ownership_share=Decimal("0.500000"),
            expected_monthly_cold_rent=Decimal("1500.00"),
            expected_monthly_utility_prepayment=Decimal("300.00"),
            yearly_non_recoverable_costs=Decimal("2400.00"),
        )
        scenario = PotentialFinancingScenario.objects.create(deal=deal, name="Bank A", owner_cash_out=Decimal("50000.00"), loan_amount=Decimal("200000.00"), interest_rate=Decimal("0.020000"), monthly_payment=Decimal("900.00"))

        metrics = potential_deal_metrics(deal, scenario)

        self.assertEqual(metrics.annual_cash_in, metrics.owner_annual_cold_rent)
        self.assertEqual(metrics.annual_cash_out, metrics.owner_operating_costs + metrics.interest_cost)
        self.assertEqual(metrics.liquidity_cashflow, metrics.annual_cash_in - metrics.owner_operating_costs - metrics.debt_service)
        self.assertEqual(metrics.cashflow, metrics.annual_cash_in - metrics.annual_cash_out)
        self.assertEqual(metrics.free_cashflow, metrics.cashflow - metrics.annual_equity_build)
        self.assertEqual(metrics.annual_equity_build, metrics.principal_repayment_estimate)
        self.assertEqual(metrics.annual_owner_roi, Decimal("0.1160"))
        self.assertEqual(metrics.gross_yield, Decimal("0.0600"))
        self.assertEqual(metrics.net_yield, Decimal("0.0520"))
        self.assertEqual(metrics.ltv, Decimal("0.6667"))
        self.assertEqual(metrics.dscr, Decimal("1.4444"))

    def test_potential_deal_portfolio_comparison_uses_selected_year_baseline(self):
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000.00"), utility_prepayment=Decimal("200.00"), total_rent=Decimal("1200.00"))
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000.00"))
        deal = PotentialDeal.objects.create(
            name="New House",
            purchase_price=Decimal("300000.00"),
            ownership_share=Decimal("0.500000"),
            expected_monthly_cold_rent=Decimal("1500.00"),
            expected_monthly_utility_prepayment=Decimal("300.00"),
        )
        scenario = PotentialFinancingScenario.objects.create(deal=deal, name="Bank A", owner_cash_out=Decimal("40000.00"), loan_amount=Decimal("200000.00"), interest_rate=Decimal("0.030000"), monthly_payment=Decimal("1000.00"))

        comparison = potential_deal_portfolio_comparison(deal, scenario, 2026)

        self.assertEqual(comparison.portfolio.year, 2026)
        self.assertEqual(comparison.delta_cashflow, Decimal("5550.00"))
        self.assertEqual(comparison.delta_monthly_cashflow, Decimal("462.50"))
        self.assertEqual(comparison.delta_debt, Decimal("100000.00"))
        self.assertEqual(comparison.delta_value, Decimal("150000.00"))
        self.assertEqual(comparison.delta_annual_rent, Decimal("9000.00"))

    def test_potential_deal_scenario_comparisons_mark_highest_annual_owner_roi(self):
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000.00"))
        deal = PotentialDeal.objects.create(
            name="New House",
            purchase_price=Decimal("300000.00"),
            ownership_share=Decimal("0.500000"),
            expected_monthly_cold_rent=Decimal("1500.00"),
            expected_monthly_utility_prepayment=Decimal("300.00"),
        )
        first = PotentialFinancingScenario.objects.create(deal=deal, name="Bank A", owner_cash_out=Decimal("40000.00"), loan_amount=Decimal("200000.00"), interest_rate=Decimal("0.030000"), monthly_payment=Decimal("1000.00"))
        PotentialFinancingScenario.objects.create(deal=deal, name="Bank B", owner_cash_out=Decimal("50000.00"), loan_amount=Decimal("180000.00"), interest_rate=Decimal("0.025000"), monthly_payment=Decimal("900.00"))

        rows = potential_deal_scenario_comparisons(deal, 2026)
        highest_annual_owner_roi_rows = [row for row in rows if row.is_highest_annual_owner_roi]

        self.assertEqual(len(highest_annual_owner_roi_rows), 1)
        self.assertEqual(highest_annual_owner_roi_rows[0].scenario, first)

    def test_potential_deal_optimizer_generates_interpolated_candidates(self):
        deal = PotentialDeal.objects.create(
            name="New House",
            purchase_price=Decimal("300000.00"),
            ownership_share=Decimal("0.500000"),
            expected_monthly_cold_rent=Decimal("2500.00"),
            yearly_non_recoverable_costs=Decimal("3000.00"),
        )

        result = optimize_potential_deal_scenario(
            deal=deal,
            maximum_cash_out=Decimal("50000.00"),
            fixed_buying_costs=Decimal("30000.00"),
            maximum_financing_percent=Decimal("83"),
            maximum_monthly_payment=Decimal("1000.00"),
            rate_100=Decimal("0.050000"),
            rate_80=Decimal("0.040000"),
            rate_60=Decimal("0.030000"),
            rate_40=Decimal("0.020000"),
        )

        self.assertEqual([option.financing_ratio for option in result.options], [Decimal("40.00"), Decimal("45.00"), Decimal("50.00"), Decimal("55.00"), Decimal("60.00"), Decimal("65.00"), Decimal("70.00"), Decimal("75.00"), Decimal("80.00"), Decimal("83.00")])
        option_83 = result.options[-1]
        self.assertEqual(option_83.interest_rate, Decimal("0.041500"))
        self.assertEqual(option_83.loan_amount, Decimal("249000.00"))
        self.assertEqual(option_83.cash_out, Decimal("40500.00"))
        self.assertTrue(option_83.is_feasible)
        self.assertEqual(result.winner, option_83)

    def test_potential_deal_optimizer_excludes_payment_and_cash_out_violations(self):
        deal = PotentialDeal.objects.create(
            name="New House",
            purchase_price=Decimal("300000.00"),
            ownership_share=Decimal("0.500000"),
            expected_monthly_cold_rent=Decimal("2500.00"),
        )

        result = optimize_potential_deal_scenario(
            deal=deal,
            maximum_cash_out=Decimal("45000.00"),
            fixed_buying_costs=Decimal("30000.00"),
            maximum_financing_percent=Decimal("90"),
            maximum_monthly_payment=Decimal("820.00"),
            rate_100=Decimal("0.050000"),
            rate_80=Decimal("0.040000"),
            rate_60=Decimal("0.030000"),
            rate_40=Decimal("0.020000"),
        )

        reasons = {option.financing_ratio: option.reason for option in result.options}
        self.assertEqual(reasons[Decimal("40.00")], "Above cash invested cap")
        self.assertEqual(reasons[Decimal("90.00")], "Interest exceeds payment cap")
        self.assertTrue(result.winner.is_feasible)
        self.assertLessEqual(result.winner.cash_out, Decimal("45000.00"))
        self.assertLessEqual(result.winner.monthly_interest_cost, Decimal("820.00"))

    def test_potential_deal_optimizer_form_requires_core_caps(self):
        form = PotentialDealOptimizerForm(
            data={
                "maximum_cash_out": "",
                "fixed_buying_costs": "0",
                "maximum_financing_percent": "100",
                "maximum_monthly_payment": "",
                "rate_100": "5",
                "rate_80": "4",
                "rate_60": "3",
                "rate_40": "2",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("maximum_cash_out", form.errors)
        self.assertIn("maximum_monthly_payment", form.errors)

    def test_potential_deal_optimizer_form_generates_missing_rate_anchors(self):
        form = PotentialDealOptimizerForm(
            data={
                "maximum_cash_out": "50000.00",
                "fixed_buying_costs": "0",
                "maximum_financing_percent": "100",
                "maximum_monthly_payment": "1500.00",
                "rate_100": "5",
                "rate_80": "",
                "rate_60": "",
                "rate_40": "",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["rate_100"], Decimal("0.050000"))
        self.assertEqual(form.cleaned_data["rate_80"], Decimal("0.044000"))
        self.assertEqual(form.cleaned_data["rate_60"], Decimal("0.039500"))
        self.assertEqual(form.cleaned_data["rate_40"], Decimal("0.038500"))
        self.assertEqual(form.generated_rate_fields, ["rate_80", "rate_60", "rate_40"])

    def test_simple_first_creation_pages_render_advanced_sections(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        property_response = self.client.get(reverse("property_create"))
        unit_response = self.client.get(reverse("unit_create"), {"property": self.property.pk})
        deal_response = self.client.get(reverse("potential_deal_create"))

        self.assertContains(property_response, "Property basics")
        self.assertContains(property_response, "Advanced property details")
        self.assertContains(property_response, "Your Cash Invested")
        self.assertContains(property_response, "Total property area")
        self.assertContains(unit_response, "Unit basics")
        self.assertContains(unit_response, "Current rent")
        self.assertContains(unit_response, "Advanced unit details")
        self.assertContains(unit_response, "Management contact")
        self.assertContains(deal_response, "Deal basics")
        self.assertContains(deal_response, "Advanced deal assumptions")
        self.assertContains(deal_response, "Your Cash Invested")
        self.assertContains(deal_response, "Financing Optimizer")

    def test_potential_deal_views_render_and_switch_scenarios(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("250000.00"))
        deal = PotentialDeal.objects.create(
            name="New House",
            address="Deal Street 1",
            purchase_price=Decimal("300000.00"),
            ownership_share=Decimal("0.500000"),
            expected_monthly_cold_rent=Decimal("1500.00"),
            expected_monthly_utility_prepayment=Decimal("300.00"),
        )
        default = PotentialFinancingScenario.objects.create(deal=deal, name="Bank A", owner_cash_out=Decimal("40000.00"), loan_amount=Decimal("200000.00"), interest_rate=Decimal("0.030000"), monthly_payment=Decimal("1000.00"), is_default=True)
        other = PotentialFinancingScenario.objects.create(deal=deal, name="Bank B", owner_cash_out=Decimal("50000.00"), loan_amount=Decimal("180000.00"), interest_rate=Decimal("0.025000"), monthly_payment=Decimal("900.00"))

        list_response = self.client.get(reverse("potential_deal_list"), {"year": "2026"})
        detail_response = self.client.get(reverse("potential_deal_detail", kwargs={"pk": deal.pk}), {"year": "2026", "scenario": str(other.pk)})
        edit_response = self.client.get(reverse("potential_deal_update", kwargs={"pk": deal.pk}))

        self.assertContains(list_response, "New House")
        self.assertContains(list_response, "Deals")
        self.assertContains(edit_response, "Showing full property values")
        self.assertContains(detail_response, "Bank A")
        self.assertContains(detail_response, "Bank B")
        self.assertContains(detail_response, "(2.50% interest)")
        self.assertContains(detail_response, f'value="{other.pk}" selected')
        self.assertContains(detail_response, reverse("potential_scenario_update", kwargs={"pk": default.pk}))
        self.assertContains(detail_response, "Selected scenario")
        self.assertContains(detail_response, 'class="selected-scenario-control"')
        self.assertContains(detail_response, f'href="{reverse("potential_deal_detail", kwargs={"pk": deal.pk})}?year=2026&scenario={other.pk}"')
        self.assertContains(detail_response, "scenario-edit-link")
        self.assertNotContains(detail_response, ">Select</a>")
        self.assertContains(detail_response, "Decision Dashboard")
        self.assertContains(detail_response, "Scenario Comparison")
        self.assertContains(detail_response, "Showing your share")
        self.assertContains(detail_response, "Showing full property values")
        self.assertContains(detail_response, "Cashflow")
        self.assertContains(detail_response, "Annual Owner ROI")
        self.assertContains(detail_response, "Your Cash Invested")
        self.assertIn("Cashflow", detail_response.content.decode())
        self.assertLess(detail_response.content.decode().index("Annual Owner ROI"), detail_response.content.decode().index("Portfolio Value"))
        self.assertLess(detail_response.content.decode().index("Portfolio Value"), detail_response.content.decode().index("Cashflow"))
        self.assertLess(detail_response.content.decode().index("Cashflow"), detail_response.content.decode().index("Debt"))
        self.assertContains(detail_response, "scenario-bar-value")
        self.assertNotContains(detail_response, "scenario-card-kpis")
        self.assertNotContains(detail_response, "Loan Payment")
        self.assertContains(detail_response, 'data-period-toggle="monthly"')
        self.assertContains(detail_response, 'data-period-toggle="annual"')
        self.assertNotContains(detail_response, "decision-kpi")
        self.assertContains(detail_response, "impact-bar")
        self.assertContains(detail_response, "scenario-card")
        self.assertContains(detail_response, "is-highest-annual-owner-roi")
        self.assertContains(detail_response, "return-stack")
        self.assertContains(detail_response, "delta-positive")
        self.assertContains(detail_response, "delta-negative")

    def test_potential_deal_optimizer_view_saves_winning_scenario(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        deal = PotentialDeal.objects.create(
            name="New House",
            purchase_price=Decimal("300000.00"),
            ownership_share=Decimal("0.500000"),
            expected_monthly_cold_rent=Decimal("2500.00"),
            yearly_non_recoverable_costs=Decimal("3000.00"),
        )

        get_response = self.client.get(reverse("potential_deal_optimizer", kwargs={"pk": deal.pk}))
        optimize_response = self.client.post(
            reverse("potential_deal_optimizer", kwargs={"pk": deal.pk}),
            {
                "action": "optimize",
                "maximum_cash_out": "50000.00",
                "fixed_buying_costs": "30000.00",
                "maximum_financing_percent": "83",
                "maximum_monthly_payment": "1000.00",
                "rate_100": "5",
                "rate_80": "4",
                "rate_60": "3",
                "rate_40": "2",
                "scenario_name": "",
            },
        )
        save_response = self.client.post(
            reverse("potential_deal_optimizer", kwargs={"pk": deal.pk}),
            {
                "action": "save",
                "maximum_cash_out": "50000.00",
                "fixed_buying_costs": "30000.00",
                "maximum_financing_percent": "83",
                "maximum_monthly_payment": "1000.00",
                "rate_100": "5",
                "rate_80": "4",
                "rate_60": "3",
                "rate_40": "2",
                "scenario_name": "Optimized Bank",
            },
        )

        self.assertContains(get_response, "Optimize Scenario")
        self.assertContains(get_response, "Showing full property values")
        self.assertContains(optimize_response, "Optimized Result")
        self.assertContains(optimize_response, "Showing your share")
        self.assertContains(optimize_response, "Showing full property values")
        self.assertContains(optimize_response, "Tested Options")
        self.assertContains(optimize_response, "Rates used")
        self.assertContains(optimize_response, "Use cold rent")
        scenario = PotentialFinancingScenario.objects.get(name="Optimized Bank")
        self.assertEqual(scenario.deal, deal)
        self.assertEqual(scenario.loan_amount, Decimal("249000.00"))
        self.assertEqual(scenario.owner_cash_out, Decimal("40500.00"))
        self.assertEqual(scenario.interest_rate, Decimal("0.041500"))
        self.assertEqual(scenario.monthly_payment, Decimal("1000.00"))
        self.assertIn("Generated by Scenario Optimizer", scenario.notes)
        self.assertRedirects(save_response, f"{reverse('potential_deal_detail', kwargs={'pk': deal.pk})}?scenario={scenario.pk}")

    def test_potential_scenario_edit_links_to_delete_confirmation(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        deal = PotentialDeal.objects.create(name="New House", purchase_price=Decimal("300000.00"))
        scenario = PotentialFinancingScenario.objects.create(deal=deal, name="Bank A", owner_cash_out=Decimal("40000.00"))

        response = self.client.get(reverse("potential_scenario_update", kwargs={"pk": scenario.pk}))

        self.assertContains(response, "Delete Scenario")
        self.assertContains(response, "Showing your share")
        self.assertContains(response, "Showing full property values")
        self.assertContains(response, reverse("potential_scenario_delete", kwargs={"pk": scenario.pk}))

    def test_potential_scenario_delete_requires_confirmation_and_removes_only_scenario(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        deal = PotentialDeal.objects.create(name="New House", purchase_price=Decimal("300000.00"))
        delete_scenario = PotentialFinancingScenario.objects.create(deal=deal, name="Bank A", owner_cash_out=Decimal("40000.00"), is_default=True)
        keep_scenario = PotentialFinancingScenario.objects.create(deal=deal, name="Bank B", owner_cash_out=Decimal("50000.00"))

        confirm = self.client.get(reverse("potential_scenario_delete", kwargs={"pk": delete_scenario.pk}))
        response = self.client.post(reverse("potential_scenario_delete", kwargs={"pk": delete_scenario.pk}))

        self.assertContains(confirm, "Delete Scenario")
        self.assertContains(confirm, "permanently deletes")
        self.assertRedirects(response, reverse("potential_deal_detail", kwargs={"pk": deal.pk}))
        self.assertFalse(PotentialFinancingScenario.objects.filter(pk=delete_scenario.pk).exists())
        self.assertTrue(PotentialFinancingScenario.objects.filter(pk=keep_scenario.pk).exists())
        keep_scenario.refresh_from_db()
        self.assertTrue(keep_scenario.is_default)

    def test_dashboard_includes_portfolio_trend_chart_data(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2025, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("240000.00"))
        AnnualPropertySnapshot.objects.create(property=self.property, year=2025, property_value=Decimal("220000.00"))
        loan = Loan.objects.create(property=self.property, name="Main loan", original_amount=Decimal("100000.00"))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2025, opening_balance=Decimal("100000.00"), closing_balance=Decimal("95000.00"), interest_paid=Decimal("2000.00"), principal_paid=Decimal("5000.00"))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2026, opening_balance=Decimal("95000.00"), closing_balance=Decimal("90000.00"), interest_paid=Decimal("1800.00"), principal_paid=Decimal("5000.00"))

        response = self.client.get(reverse("dashboard"))

        self.assertContains(response, "Portfolio Trend")
        self.assertContains(response, "Data quality for 2026")
        self.assertContains(response, "Non-recoverable costs missing for Test House. Using 5% rent estimate.")
        self.assertNotContains(response, "Rent history complete")
        self.assertNotContains(response, "Loan balance entered")
        self.assertContains(response, "Showing your share")
        self.assertContains(response, 'data-trend-metric="cashflow"')
        chart_data = response.context["chart_data"]
        self.assertEqual(chart_data["trend"]["labels"], [2025, 2026])
        expected_2025 = portfolio_performance(2025)
        expected_2026 = portfolio_performance(2026)
        self.assertEqual(chart_data["trend"]["metrics"]["cashflow"], [float(expected_2025.total_cashflow), float(expected_2026.total_cashflow)])
        self.assertEqual(chart_data["trend"]["metrics"]["debt"], [float(expected_2025.total_debt), float(expected_2026.total_debt)])
        self.assertEqual(chart_data["trend"]["metrics"]["ltv"], [float(expected_2025.ltv * 100), float(expected_2026.ltv * 100)])
        self.assertEqual(chart_data["trend"]["metrics"]["value"], [float(expected_2025.total_value), float(expected_2026.total_value)])
        self.assertEqual(chart_data["trend"]["breakdowns"]["cashflow"][0][0]["label"], "Test House")
        self.assertEqual(chart_data["trend"]["breakdowns"]["cashflow"][0][0]["value"], float(expected_2025.rows[0].cashflow))
        self.assertEqual(chart_data["trend"]["breakdowns"]["value"][0][0]["value"], float(expected_2025.rows[0].property_value))
        self.assertEqual(chart_data["trend"]["breakdowns"]["debt"][0][0]["label"], "Test House - Main loan")
        self.assertEqual(chart_data["trend"]["breakdowns"]["debt"][0][0]["value"], float(expected_2025.rows[0].closing_debt))
        self.assertEqual(chart_data["trend"]["breakdowns"]["ltv"][0][0]["debt"], float(expected_2025.rows[0].closing_debt))
        self.assertEqual(chart_data["trend"]["breakdowns"]["ltv"][0][0]["value"], float(expected_2025.rows[0].property_value))

    def test_dashboard_defaults_to_after_tax_and_can_switch_to_before_tax(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("240000.00"))

        default_response = self.client.get(reverse("dashboard"), {"year": "2026"})
        before_response = self.client.get(reverse("dashboard"), {"year": "2026", "tax": "before"})

        self.assertEqual(default_response.context["tax_mode"], "after")
        self.assertContains(default_response, "After tax estimate")
        self.assertContains(default_response, "Portfolio Tax Bridge")
        self.assertContains(default_response, "Cashflow before tax")
        self.assertContains(default_response, "Taxable result")
        self.assertEqual(before_response.context["tax_mode"], "before")
        self.assertContains(before_response, "Before tax")
        self.assertContains(before_response, 'value="before" checked')

    def test_dashboard_hides_tax_ui_when_tax_is_disabled(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        AppSettings.objects.create(pk=1, tax_calculations_enabled=False, effective_tax_rate=Decimal("0.250000"), tax_loss_benefit_enabled=True)
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("240000.00"))

        response = self.client.get(reverse("dashboard"), {"year": "2026", "tax": "after"})

        self.assertEqual(response.context["tax_mode"], "before")
        self.assertFalse(response.context["tax_calculations_enabled"])
        self.assertNotContains(response, "After tax estimate")
        self.assertNotContains(response, "Before Tax")
        self.assertNotContains(response, "Portfolio Tax Bridge")
        self.assertEqual(response.context["portfolio"].total_cashflow, response.context["portfolio"].total_cashflow_before_tax)

    def test_dashboard_after_tax_chart_data_uses_selected_mode(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        AppSettings.objects.create(pk=1, effective_tax_rate=Decimal("0.250000"), tax_loss_benefit_enabled=True)
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("240000.00"))

        after_response = self.client.get(reverse("dashboard"), {"year": "2026"})
        before_response = self.client.get(reverse("dashboard"), {"year": "2026", "tax": "before"})
        after_expected = portfolio_performance(2026, tax_mode="after", tax_settings=AppSettings.load())
        before_expected = portfolio_performance(2026, tax_mode="before", tax_settings=AppSettings.load())

        self.assertEqual(after_response.context["chart_data"]["trend"]["metrics"]["cashflow"], [float(after_expected.total_cashflow)])
        self.assertEqual(before_response.context["chart_data"]["trend"]["metrics"]["cashflow"], [float(before_expected.total_cashflow)])
        self.assertNotEqual(after_response.context["portfolio"].total_cashflow, before_response.context["portfolio"].total_cashflow)

    def test_dashboard_trend_includes_property_after_backfilled_purchase_value_years(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        other_property = Property.objects.create(
            name="Earlier Property",
            address="Earlier Street 1",
            ownership_share=Decimal("0.5"),
            purchase_price=Decimal("300000.00"),
            acquisition_date=date(2021, 12, 1),
        )
        AnnualPropertySnapshot.objects.create(property=other_property, year=2026, property_value=Decimal("350000.00"))
        loan = Loan.objects.create(property=other_property, name="Earlier Property loan", original_amount=Decimal("250000.00"), start_date=date(2021, 12, 1))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2021, opening_balance=Decimal("250000.00"), closing_balance=Decimal("245000.00"))

        backfill_property_snapshots(other_property, through_year=2026)
        response = self.client.get(reverse("dashboard"))

        chart_data = response.context["chart_data"]
        year_index = chart_data["trend"]["labels"].index(2021)
        self.assertGreater(chart_data["trend"]["metrics"]["value"][year_index], 0)
        self.assertGreater(chart_data["trend"]["metrics"]["debt"][year_index], 0)

    def test_dashboard_property_filter_limits_portfolio_data(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("200000.00"))
        other_property = Property.objects.create(
            name="Filtered House",
            address="Filter Street 1",
            ownership_share=Decimal("0.5"),
            purchase_price=Decimal("400000.00"),
            cash_invested_at_purchase=Decimal("50000.00"),
        )
        AnnualPropertySnapshot.objects.create(property=other_property, year=2026, property_value=Decimal("400000.00"))

        response = self.client.get(reverse("dashboard"), {"year": "2026", "properties": str(other_property.pk)})

        self.assertContains(response, "Filtered House")
        self.assertContains(response, "All properties")
        self.assertEqual([row.property.pk for row in response.context["portfolio"].rows], [other_property.pk])
        self.assertEqual(response.context["portfolio"].total_value, Decimal("200000.00"))
        self.assertEqual(response.context["chart_data"]["labels"], ["Filtered House"])
        quality_messages = [item.message for item in response.context["data_quality"].items]
        self.assertIn("Non-recoverable costs missing for Filtered House. Using 5% rent estimate.", quality_messages)
        self.assertNotIn("Non-recoverable costs missing for Test House. Using 5% rent estimate.", quality_messages)
        self.assertContains(response, "Data quality for 2026")

    def test_dashboard_hides_data_quality_panel_when_inputs_are_complete(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        snapshot = AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("200000.00"))
        AnnualPropertyCost.objects.create(snapshot=snapshot, category=AnnualPropertyCost.OTHER, amount=Decimal("1000.00"), notes="Yearly non-recoverable costs")
        loan = Loan.objects.create(property=self.property, name="Main loan", original_amount=Decimal("100000.00"))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2026, opening_balance=Decimal("100000.00"), closing_balance=Decimal("95000.00"))

        response = self.client.get(reverse("dashboard"), {"year": "2026"})

        self.assertNotContains(response, "Data quality for 2026")
        self.assertEqual(response.context["data_quality_warnings"], [])

    def test_property_list_renders_property_work_cards(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        RentPeriod.objects.create(lease=self.lease, effective_start=date.today(), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))

        response = self.client.get(reverse("property_list"))

        self.assertContains(response, "The main working area")
        self.assertNotContains(response, "Current Rent")
        self.assertNotContains(response, "Annualized Rent")
        self.assertNotContains(response, "Annual ROI")
        self.assertContains(response, f'href="{reverse("property_detail", kwargs={"pk": self.property.pk})}"')
        self.assertNotContains(response, f'href="{reverse("property_detail", kwargs={"pk": self.property.pk})}#units"')
        self.assertNotContains(response, f'href="{reverse("property_detail", kwargs={"pk": self.property.pk})}#loans"')
        self.assertContains(response, "1 unit")
        self.assertNotContains(response, "1 occupied")
        self.assertContains(response, "0 loans")
        self.assertNotContains(response, "1 units")
        self.assertNotContains(response, "0 loan</a>")
        self.assertNotContains(response, "Fully rented")

    def test_tenant_list_defaults_to_current_contract_tenants(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        RentPeriod.objects.create(lease=self.lease, effective_start=self.lease.start_date, cold_rent=Decimal("1000.00"), utility_prepayment=Decimal("200.00"), total_rent=Decimal("1200.00"))
        LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=self.lease.start_date)
        occupant = Tenant.objects.create(first_name="Grace", last_name="Hopper", email="grace@example.com")
        LeasePerson.objects.create(lease=self.lease, person=occupant, role=LeasePerson.OCCUPANT, move_in_date=self.lease.start_date)
        former_tenant = Tenant.objects.create(first_name="Alan", last_name="Turing")
        former_lease = Lease.objects.create(unit=self.unit, tenant=former_tenant, start_date=date(2020, 1, 1), end_date=date(2020, 12, 31), is_active=False)
        LeasePerson.objects.create(lease=former_lease, person=former_tenant, role=LeasePerson.PRIMARY, move_in_date=former_lease.start_date, move_out_date=former_lease.end_date)

        response = self.client.get(reverse("tenant_list"))

        self.assertContains(response, "Tenants")
        self.assertContains(response, "Lovelace, Ada")
        self.assertContains(response, "€1,000.00")
        self.assertContains(response, "€1,200.00")
        self.assertContains(response, "Primary contract tenant")
        self.assertNotContains(response, "Hopper, Grace")
        self.assertNotContains(response, "Turing, Alan")
        self.assertContains(response, 'option value="contract" selected')
        self.assertContains(response, 'option value="current" selected')

    def test_tenant_list_everyone_filter_includes_occupants_and_financial_support_office_search(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        self.tenant.support_office_name = "Financial Help Office"
        self.tenant.support_office_email = "support@example.com"
        self.tenant.support_office_phone = "12345"
        self.tenant.save()
        LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=self.lease.start_date)
        occupant = Tenant.objects.create(first_name="Grace", last_name="Hopper", email="grace@example.com", phone="555", support_office_email="support@example.com")
        LeasePerson.objects.create(lease=self.lease, person=occupant, role=LeasePerson.OCCUPANT, move_in_date=self.lease.start_date)

        response = self.client.get(reverse("tenant_list"), {"role": "everyone", "status": "everyone", "q": "support@example.com", "sort": "property"})

        self.assertContains(response, "Lovelace, Ada")
        self.assertContains(response, "Hopper, Grace")
        self.assertContains(response, "Financial Support Office")
        self.assertContains(response, "support@example.com")
        self.assertContains(response, "Edit person")
        self.assertContains(response, "Open unit")
        self.assertEqual(response.context["sort"], "property")

    def test_property_detail_renders_single_dossier_page(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        response = self.client.get(reverse("property_detail", kwargs={"pk": self.property.pk}))

        self.assertContains(response, "Property Info")
        self.assertContains(response, "Units")
        self.assertContains(response, "?edit=1")
        self.assertNotContains(response, "Back to properties")
        self.assertNotContains(response, "data-tab-target")
        self.assertNotContains(response, "Units &amp; Tenants")
        self.assertNotContains(response, "data-tab-panel=\"admin\"")
        self.assertNotContains(response, "data-tab-target=\"loans\"")
        self.assertNotContains(response, "data-tab-panel=\"yearly\"")
        self.assertNotContains(response, "data-tab-panel=\"exports\"")
        self.assertNotContains(response, "Annual ROI")
        self.assertNotContains(response, "Cashflow")

    def test_property_detail_inline_edit_saves_dossier_and_photo(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        with TemporaryDirectory() as media_root, override_settings(MEDIA_ROOT=media_root):
            response = self.client.post(
                reverse("property_detail", kwargs={"pk": self.property.pk}),
                {
                    "name": "Updated House",
                    "object_type": self.property.object_type,
                    "street_address": "Updated Street 2",
                    "postal_code": "34117",
                    "city": "Kassel",
                    "construction_year": "1928",
                    "total_building_area_sqm": "123.45",
                    "notes": "Dossier notes",
                    "photo": SimpleUploadedFile("house.jpg", b"fake image data", content_type="image/jpeg"),
                },
            )

            self.assertRedirects(response, reverse("property_detail", kwargs={"pk": self.property.pk}))
            self.property.refresh_from_db()
            administration = PropertyAdministration.objects.get(property=self.property)
            self.assertEqual(self.property.name, "Updated House")
            self.assertEqual(self.property.street_address, "Updated Street 2")
            self.assertEqual(self.property.address, "Updated Street 2\n34117 Kassel")
            self.assertEqual(administration.construction_year, 1928)
            self.assertEqual(administration.total_building_area_sqm, Decimal("123.45"))
            self.assertTrue(self.property.photo.name.startswith("property_photos/"))

    def test_property_notes_are_editable_without_full_edit_mode(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        response = self.client.post(
            reverse("property_detail", kwargs={"pk": self.property.pk}),
            {"form_kind": "notes", "notes": "Always editable note"},
        )

        self.assertRedirects(response, reverse("property_detail", kwargs={"pk": self.property.pk}))
        self.property.refresh_from_db()
        self.assertEqual(self.property.notes, "Always editable note")
        detail = self.client.get(reverse("property_detail", kwargs={"pk": self.property.pk}))
        self.assertContains(detail, "Save Notes")
        self.assertContains(detail, "Always editable note")
        self.assertContains(detail, 'data-autogrow')

    def test_property_detail_hides_loans_section_without_loans(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        response = self.client.get(reverse("property_detail", kwargs={"pk": self.property.pk}))

        self.assertNotContains(response, '<h2>Loans</h2>', html=True)
        self.assertNotContains(response, "loan-card")

    def test_property_detail_shows_clickable_loan_boxes(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        first_loan = Loan.objects.create(property=self.property, name="Main loan", lender="Hausbank", original_amount=Decimal("200000.00"))
        second_loan = Loan.objects.create(property=self.property, name="Modernization loan", original_amount=Decimal("25000.00"))

        response = self.client.get(reverse("property_detail", kwargs={"pk": self.property.pk}))

        self.assertContains(response, '<h2>Loans</h2>', html=True)
        self.assertContains(response, reverse("loan_update", kwargs={"pk": first_loan.pk}))
        self.assertContains(response, reverse("loan_update", kwargs={"pk": second_loan.pk}))
        self.assertContains(response, "Main loan")
        self.assertContains(response, "Hausbank")
        self.assertContains(response, "Modernization loan")
        self.assertContains(response, "No lender stored")
        self.assertContains(response, "€200,000.00")
        self.assertContains(response, "€25,000.00")

    def test_exports_page_keeps_backups_separate(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        response = self.client.get(reverse("exports"))

        self.assertContains(response, "Bank Financing Overview")
        self.assertContains(response, "Showing full property values")
        self.assertNotContains(response, "Current Rent Overview")
        self.assertNotContains(response, "Portfolio Reports")
        self.assertNotContains(response, "Database backup and restore")
        self.assertNotContains(response, "Create Database Backup")

    def test_property_yearly_edit_form_shows_full_property_badge(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        response = self.client.get(reverse("property_update", kwargs={"pk": self.property.pk}), {"year": "2026"})

        self.assertContains(response, "Annual Property History")
        self.assertContains(response, "Showing full property values")

    def test_backup_and_settings_pages_render_from_more_menu(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        backup_response = self.client.get(reverse("backup"))
        settings_response = self.client.get(reverse("settings"))

        self.assertContains(backup_response, "Create Backup")
        self.assertContains(backup_response, "Restore Database")
        self.assertContains(settings_response, "Local App Paths")
        self.assertContains(settings_response, "Tax Settings")
        self.assertNotContains(settings_response, "Landlord Settings")

    def test_settings_page_saves_tax_settings(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        response = self.client.post(
            reverse("settings"),
            {
                "tax_calculations_enabled": "on",
                "effective_tax_rate": "25",
                "tax_loss_benefit_enabled": "on",
                "tax_deductible_costs_2026": "1234.56",
                "tax_notes_2026": "Tax advisor estimate",
            },
        )

        self.assertRedirects(response, reverse("settings"))
        app_settings = AppSettings.load()
        self.assertEqual(app_settings.effective_tax_rate, Decimal("0.250000"))
        self.assertTrue(app_settings.tax_loss_benefit_enabled)
        tax_row = AnnualPortfolioTax.objects.get(year=2026)
        self.assertEqual(tax_row.tax_deductible_costs, Decimal("1234.56"))
        self.assertEqual(tax_row.notes, "Tax advisor estimate")

    def test_settings_page_can_disable_tax_calculations(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        AppSettings.objects.create(pk=1, tax_calculations_enabled=True, effective_tax_rate=Decimal("0.250000"), tax_loss_benefit_enabled=True)

        response = self.client.post(reverse("settings"), {})

        self.assertRedirects(response, reverse("settings"))
        app_settings = AppSettings.load()
        self.assertFalse(app_settings.tax_calculations_enabled)
        self.assertEqual(app_settings.effective_tax_rate, Decimal("0.250000"))

    def test_current_rent_overview_includes_active_and_vacant_units(self):
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        Unit.objects.create(property=self.property, label="Vacant Unit")

        overview = current_rent_overview(date(2026, 5, 22))

        self.assertEqual(len(overview.units), 2)
        self.assertEqual(overview.total_monthly_cold_rent, Decimal("1000.00"))
        self.assertEqual(overview.total_monthly_utility_prepayment, Decimal("200.00"))
        self.assertEqual(overview.total_monthly_total_rent, Decimal("1200.00"))
        self.assertEqual(overview.total_annual_total_rent, Decimal("14400.00"))
        self.assertEqual(overview.total_owner_annual_total_rent, Decimal("7200.00"))
        vacant = next(row for row in overview.units if row.unit_label == "Vacant Unit")
        self.assertEqual(vacant.tenant_name, "Vacant")
        self.assertEqual(vacant.monthly_total_rent, Decimal("0.00"))
    def test_property_dossier_saves_object_type(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        response = self.client.post(
            reverse("property_detail", kwargs={"pk": self.property.pk}),
            {
                "name": self.property.name,
                "object_type": Property.MULTI_FAMILY_HOUSE,
                "street_address": "Test Street 1",
                "postal_code": "34117",
                "city": "Kassel",
                "construction_year": "1930",
                "total_building_area_sqm": "250.00",
                "notes": self.property.notes,
            },
        )

        self.assertRedirects(response, reverse("property_detail", kwargs={"pk": self.property.pk}))
        self.property.refresh_from_db()
        self.assertEqual(self.property.object_type, Property.MULTI_FAMILY_HOUSE)
        detail = self.client.get(reverse("property_detail", kwargs={"pk": self.property.pk}))
        self.assertContains(detail, "Multi-family house")

    def test_bank_financing_overview_uses_current_full_property_values(self):
        self.property.object_type = Property.MULTI_FAMILY_HOUSE
        self.property.purchase_price = Decimal("200000.00")
        self.property.save()
        PropertyAdministration.objects.create(property=self.property, construction_year=1930, total_building_area_sqm=Decimal("250.00"))
        self.unit.area_sqm = Decimal("80.00")
        self.unit.save()
        Unit.objects.create(property=self.property, label="Unit 2", area_sqm=Decimal("70.00"))
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        AnnualPropertySnapshot.objects.create(property=self.property, year=2024, property_value=Decimal("230000.00"))
        AnnualPropertySnapshot.objects.create(property=self.property, year=2026, property_value=Decimal("260000.00"))
        loan = Loan.objects.create(
            property=self.property,
            name="Bank loan",
            lender="Hausbank",
            original_amount=Decimal("120000.00"),
            default_interest_rate=Decimal("0.035"),
            default_monthly_payment=Decimal("800.00"),
        )
        snapshot = AnnualLoanSnapshot.objects.create(
            loan=loan,
            year=2026,
            opening_balance=Decimal("100000.00"),
            closing_balance=Decimal("88000.00"),
            monthly_payment=Decimal("900.00"),
            interest_rate=Decimal("0.0325"),
            rate_reset_date=date(2031, 12, 31),
        )

        overview = bank_financing_overview(date(2026, 7, 2))

        self.assertEqual(len(overview.real_estate), 1)
        real_estate = overview.real_estate[0]
        self.assertEqual(real_estate.object_type, "Multi-family house")
        self.assertEqual(real_estate.living_area_sqm, Decimal("150"))
        self.assertEqual(real_estate.total_area_sqm, Decimal("250.00"))
        self.assertEqual(real_estate.property_value, Decimal("260000.00"))
        self.assertEqual(real_estate.annual_cold_rent, Decimal("12000.00"))
        self.assertFalse(real_estate.uses_purchase_price_fallback)
        self.assertFalse(real_estate.uses_total_area_fallback)
        self.assertFalse(real_estate.has_no_area_fallback)
        loan_row = overview.loans[0]
        self.assertEqual(loan_row.initial_loan_value, Decimal("120000.00"))
        self.assertEqual(loan_row.lender, "Hausbank")
        self.assertEqual(loan_row.current_loan_amount, current_debt_from_annual_snapshot(snapshot.opening_balance, snapshot.closing_balance, 2026, date(2026, 7, 2)))
        self.assertEqual(loan_row.fixed_until_year, 2031)
        self.assertEqual(loan_row.monthly_payment, Decimal("900.00"))
        self.assertEqual(loan_row.interest_rate, Decimal("0.0325"))

    def test_bank_financing_preview_aggregates_property_rows_and_warnings(self):
        self.property.purchase_price = Decimal("200000.00")
        self.property.save()
        PropertyAdministration.objects.create(property=self.property, total_building_area_sqm=Decimal("90.00"))
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        first_loan = Loan.objects.create(property=self.property, name="Loan A", original_amount=Decimal("50000.00"), default_interest_rate=Decimal("0.040000"), default_monthly_payment=Decimal("500.00"))
        second_loan = Loan.objects.create(property=self.property, name="Loan B", original_amount=Decimal("100000.00"), default_interest_rate=Decimal("0.020000"), default_monthly_payment=Decimal("750.00"))
        AnnualLoanSnapshot.objects.create(loan=first_loan, year=2026, opening_balance=Decimal("50000.00"), closing_balance=Decimal("50000.00"), monthly_payment=Decimal("500.00"), interest_rate=Decimal("0.040000"), rate_reset_date=date(2030, 12, 31))
        AnnualLoanSnapshot.objects.create(loan=second_loan, year=2026, opening_balance=Decimal("100000.00"), closing_balance=Decimal("100000.00"), monthly_payment=Decimal("750.00"), interest_rate=Decimal("0.020000"), rate_reset_date=date(2034, 12, 31))

        preview = bank_financing_preview(date(2026, 1, 1))

        self.assertEqual(len(preview.rows), 1)
        row = preview.rows[0]
        self.assertEqual(row.property, self.property)
        self.assertEqual(row.property_value, Decimal("200000.00"))
        self.assertEqual(row.annual_cold_rent, Decimal("12000.00"))
        self.assertEqual(row.debt, Decimal("150000.00"))
        self.assertEqual(row.monthly_payment, Decimal("1250.00"))
        self.assertEqual(row.interest_rate, Decimal("0.02666666666666666666666666667"))
        self.assertEqual(row.fixed_until_year, 2030)
        self.assertIn("Property value missing for Test House. Purchase price will be used.", preview.warnings)
        self.assertIn("Living area missing for Test House. Total property area will be used instead.", preview.warnings)

    def test_bank_financing_export_workbook_and_endpoints(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        self.property.object_type = Property.APARTMENT
        self.property.save()
        PropertyAdministration.objects.create(property=self.property, construction_year=1980, total_building_area_sqm=Decimal("90.00"))
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        Loan.objects.create(property=self.property, name="Fallback loan", original_amount=Decimal("50000.00"))

        workbook_bytes = export_bank_financing_workbook(date(2026, 5, 22))
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
        pdf_response = self.client.get(reverse("export_bank_financing_pdf"))
        excel_response = self.client.get(reverse("export_bank_financing_excel"))

        self.assertEqual(workbook.sheetnames, ["Real Estate", "Loans"])
        self.assertEqual(workbook["Real Estate"]["F3"].value, "Living area sqm")
        self.assertEqual(workbook["Real Estate"]["G3"].value, "Total property area sqm")
        self.assertEqual(workbook["Real Estate"]["A4"].value, "Test House")
        self.assertEqual(workbook["Real Estate"]["B4"].value, "Apartment")
        self.assertEqual(workbook["Real Estate"]["I4"].value, 12000)
        self.assertEqual(workbook["Loans"]["A4"].value, "Test House")
        self.assertEqual(workbook["Loans"]["C4"].value, 50000)
        self.assertIsNone(workbook["Loans"]["F4"].value)
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response["Content-Type"], "application/pdf")
        self.assertEqual(excel_response.status_code, 200)
        self.assertEqual(excel_response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertTrue(ReportExport.objects.filter(export_type="bank_financing_excel").exists())

        exports_response = self.client.get(reverse("exports"))
        self.assertContains(exports_response, "Bank Financing Export Preview")
        self.assertContains(exports_response, "Showing full property values")
        self.assertContains(exports_response, "Property")
        self.assertContains(exports_response, "Monthly payment")
        self.assertContains(exports_response, "Property value missing for Test House. Purchase price will be used.")
        self.assertContains(exports_response, "Living area missing for Test House. Total property area will be used instead.")

    def test_bank_financing_export_save_flow_creates_openable_file(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        with TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                response = self.client.get(reverse("export_bank_financing_excel"), {"save": "1"})
                file_name = ReportExport.objects.filter(export_type="bank_financing_excel").latest("created_at").file_name
                file_response = self.client.get(reverse("export_file", kwargs={"file_name": file_name}))

                self.assertRedirects(response, reverse("exports"))
                self.assertTrue((Path(media_root) / "exports" / file_name).exists())
                self.assertEqual(file_response.status_code, 200)
                self.assertIn("inline", file_response["Content-Disposition"])

    def test_unit_administration_form_saves_typed_admin_data_and_contact(self):
        form = UnitAdministrationForm(
            data={
                "postal_code": "34119",
                "construction_year": "1928",
                "total_building_area_sqm": "10422.40",
                "property_admin_notes": "Building note",
                "apartment_number": "31",
                "cellar_number": "K1",
                "ownership_share_text": "48/10.000",
                "monthly_house_fee": "290.00",
                "unit_admin_notes": "Unit note",
                "local_court": "Kassel",
                "land_register_district": "Wehlheiden",
                "sheet_number": "6104",
                "cadastral_district": "2",
                "plot_numbers": "20/35",
                "land_registry_notes": "",
                "heating_type": "Fernwärme",
                "boiler_installation_info": "-",
                "instant_water_heater_info": "",
                "technical_notes": "",
                "contact_mode": "new",
                "existing_contact": "",
                "contact_name": "Hausverwaltung ImmoPower",
                "contact_email": "info@example.com",
                "contact_phone": "123",
                "contact_notes": "",
            },
            unit=self.unit,
        )

        self.assertTrue(form.is_valid(), form.errors)
        form.save()

        self.assertEqual(PropertyAdministration.objects.get(property=self.property).construction_year, 1928)
        self.assertEqual(UnitAdministration.objects.get(unit=self.unit).apartment_number, "31")
        self.assertEqual(UnitLandRegistry.objects.get(unit=self.unit).local_court, "Kassel")
        self.assertEqual(UnitTechnicalInfo.objects.get(unit=self.unit).heating_type, "Fernwärme")
        link = UnitContact.objects.select_related("contact").get(unit=self.unit)
        self.assertEqual(link.contact.name, "Hausverwaltung ImmoPower")
        self.assertEqual(link.contact.email, "info@example.com")

    def test_unit_administration_can_reuse_contact_across_units(self):
        contact = Contact.objects.create(name="Freiraum Immobilien", email="info@example.com")
        other_unit = Unit.objects.create(property=self.property, label="Unit 2")
        for unit in (self.unit, other_unit):
            form = UnitAdministrationForm(
                data={
                    "postal_code": "",
                    "construction_year": "",
                    "total_building_area_sqm": "",
                    "property_admin_notes": "",
                    "apartment_number": "",
                    "cellar_number": "",
                    "ownership_share_text": "",
                    "monthly_house_fee": "",
                    "unit_admin_notes": "",
                    "local_court": "",
                    "land_register_district": "",
                    "sheet_number": "",
                    "cadastral_district": "",
                    "plot_numbers": "",
                    "land_registry_notes": "",
                    "heating_type": "",
                    "boiler_installation_info": "",
                    "instant_water_heater_info": "",
                    "technical_notes": "",
                    "contact_mode": "existing",
                    "existing_contact": str(contact.pk),
                    "contact_name": "",
                    "contact_email": "",
                    "contact_phone": "",
                    "contact_notes": "",
                },
                unit=unit,
            )
            self.assertTrue(form.is_valid(), form.errors)
            form.save()

        self.assertEqual(UnitContact.objects.filter(contact=contact).count(), 2)

    def test_unit_detail_renders_unit_and_technical_info(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        UnitAdministration.objects.create(unit=self.unit, apartment_number="31", ownership_share_text="48/10.000", monthly_house_fee=Decimal("290.00"))
        UnitLandRegistry.objects.create(unit=self.unit, local_court="Kassel", land_register_district="Wehlheiden", sheet_number="6104")
        UnitTechnicalInfo.objects.create(unit=self.unit, heating_type="Fernwärme")
        contact = Contact.objects.create(name="Hausverwaltung ImmoPower", email="info@example.com")
        UnitContact.objects.create(unit=self.unit, contact=contact)

        response = self.client.get(reverse("unit_detail", kwargs={"pk": self.unit.pk}))

        self.assertContains(response, "Unit Info")
        self.assertContains(response, "Technical Info")
        self.assertContains(response, "Fernwärme")
        self.assertNotContains(response, "data-tab-target")
        self.assertNotContains(response, "Administration")

    def test_unit_detail_shows_contract_tenants_occupants_and_rent_history(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=date(2026, 1, 1))
        child = Tenant.objects.create(first_name="Mini", last_name="Lovelace")
        LeasePerson.objects.create(lease=self.lease, person=child, role=LeasePerson.CHILD, move_in_date=date(2026, 1, 1))
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))

        response = self.client.get(reverse("unit_detail", kwargs={"pk": self.unit.pk}))

        self.assertContains(response, "Contract Tenants")
        self.assertContains(response, "Other Occupants")
        self.assertContains(response, "Lovelace, Ada")
        self.assertContains(response, "Lovelace, Mini")
        self.assertContains(response, "Rent valid from")
        self.assertContains(response, "Corrections")
        self.assertContains(response, reverse("lease_people_update", kwargs={"pk": self.lease.pk}))

    def test_unit_detail_renders_mietbescheinigung_section(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)

        response = self.client.get(reverse("unit_detail", kwargs={"pk": self.unit.pk}))

        self.assertContains(response, "Mietbescheinigung")
        self.assertContains(response, "Jobcenter")
        self.assertContains(response, "Stadt Kassel")
        self.assertContains(response, reverse("unit_mietbescheinigung", kwargs={"pk": self.unit.pk}))

    def test_mietbescheinigung_form_prefills_contract_tenant_and_rent(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        self.property.street_address = "Test Street 1"
        self.property.postal_code = "34117"
        self.property.city = "Kassel"
        self.property.address = "Legacy should not be used"
        self.property.save()
        landlord = LandlordProfile.objects.create(
            name="Casa Owner",
            street_address="Owner Street 2",
            postal_code="34119",
            city="Kassel",
            email="owner@example.com",
            is_default=True,
        )
        LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=date(2026, 1, 1))
        occupant = Tenant.objects.create(first_name="Mini", last_name="Lovelace")
        LeasePerson.objects.create(lease=self.lease, person=occupant, role=LeasePerson.CHILD, move_in_date=date(2026, 1, 1))
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))

        response = self.client.get(reverse("unit_mietbescheinigung", kwargs={"pk": self.unit.pk}), {"template": "stadt_kassel", "landlord": landlord.pk})

        self.assertContains(response, "Create Mietbescheinigung")
        self.assertContains(response, f'action="{reverse("unit_mietbescheinigung", kwargs={"pk": self.unit.pk})}"')
        self.assertContains(response, "URL.revokeObjectURL(url);")
        self.assertContains(response, "60000")
        self.assertContains(response, "New landlord")
        self.assertContains(response, "Casa Owner")
        self.assertContains(response, "Test Street 1")
        self.assertContains(response, "34117")
        self.assertContains(response, "Kassel")
        self.assertContains(response, "Lovelace, Ada")
        self.assertContains(response, 'value="1"')
        self.assertContains(response, 'value="1000.00"')
        self.assertNotContains(response, "Lovelace, Mini")

    def test_mietbescheinigung_generation_saves_pdf_export(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        self.property.street_address = "Test Street 1"
        self.property.postal_code = "34117"
        self.property.city = "Kassel"
        self.property.address = "Test Street 1\n34117 Kassel"
        self.property.save()
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=date(2026, 1, 1))
        post_data = {
            "template": "jobcenter",
            "landlord_profile": "__new__",
            "landlord_name": "Casa Owner",
            "landlord_street": "Owner Street 2",
            "landlord_zip": "34119",
            "landlord_city": "Kassel",
            "landlord_phone": "123",
            "landlord_fax": "",
            "landlord_email": "owner@example.com",
            "tenant_name": "Lovelace, Ada",
            "tenant_street": "Test Street 1",
            "tenant_zip": "34117",
            "tenant_city": "Kassel",
            "tenant_contact": "tenant@example.com",
            "lease_start": "01.01.2026",
            "tenant_count": "1",
            "floor": "1",
            "construction_year": "",
            "building_area_sqm": "",
            "living_area_sqm": "55.00",
            "rooms": "",
            "rent_valid_from": "01.01.2026",
            "cold_rent": "1000.00",
            "operating_costs": "200.00",
            "total_rent": "1200.00",
            "heating_costs": "",
            "warm_water_costs": "",
            "garage_cost": "",
            "parking_cost": "",
            "arrears_amount": "",
            "operating_costs_advance": "on",
            "issue_place": "Kassel",
            "issue_date": "01.06.2026",
        }

        with TemporaryDirectory() as media_root, override_settings(MEDIA_ROOT=media_root):
            for template in ("jobcenter", "stadt_kassel"):
                response = self.client.post(reverse("unit_mietbescheinigung", kwargs={"pk": self.unit.pk}), {**post_data, "template": template})

                self.assertEqual(response.status_code, 200)
                self.assertEqual(response["Content-Type"], "application/pdf")
                self.assertIn("attachment;", response["Content-Disposition"])
                self.assertTrue(response.content.startswith(b"%PDF"))
                export = ReportExport.objects.get(export_type="mietbescheinigung_pdf", file_name__contains=template)
                path = Path(media_root) / "exports" / export.file_name
                self.assertTrue(path.exists())
                self.assertTrue(path.read_bytes().startswith(b"%PDF"))
                self.assertEqual(export.property, self.property)

    def test_mietbescheinigung_save_new_landlord_with_signature(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=date(2026, 1, 1))
        post_data = {
            "action": "save_landlord",
            "template": "jobcenter",
            "landlord_profile": "__new__",
            "landlord_name": "New Owner",
            "landlord_street": "Owner Street 2",
            "landlord_zip": "34119",
            "landlord_city": "Kassel",
            "landlord_phone": "123",
            "landlord_fax": "",
            "landlord_email": "owner@example.com",
            "tenant_name": "Lovelace, Ada",
            "tenant_street": "Test Street 1",
            "tenant_zip": "34117",
            "tenant_city": "Kassel",
            "lease_start": "01.01.2026",
            "tenant_count": "1",
            "rent_valid_from": "01.01.2026",
            "cold_rent": "1000.00",
            "operating_costs": "200.00",
            "total_rent": "1200.00",
            "issue_date": "01.06.2026",
        }

        with TemporaryDirectory() as media_root, override_settings(MEDIA_ROOT=media_root):
            response = self.client.post(
                reverse("unit_mietbescheinigung", kwargs={"pk": self.unit.pk}),
                {**post_data, "signature_image": SimpleUploadedFile("signature.png", SIGNATURE_PNG, content_type="image/png")},
            )

            landlord = LandlordProfile.objects.get(name="New Owner")
            self.assertRedirects(response, f"{reverse('unit_mietbescheinigung', kwargs={'pk': self.unit.pk})}?template=jobcenter&landlord={landlord.pk}")
            self.assertTrue(landlord.signature_image.name.startswith("landlord_signatures/"))
            self.assertTrue((Path(media_root) / landlord.signature_image.name).exists())

    def test_mietbescheinigung_generation_blocks_missing_active_lease(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        self.lease.is_active = False
        self.lease.save()

        response = self.client.get(reverse("unit_mietbescheinigung", kwargs={"pk": self.unit.pk}))

        self.assertContains(response, "This unit has no active tenant period.")
        self.assertContains(response, "disabled")

    def test_unit_inline_edit_saves_unit_technical_people_and_dated_rent_change(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        link = LeasePerson.objects.create(lease=self.lease, person=self.tenant, role=LeasePerson.PRIMARY, move_in_date=date(2026, 1, 1))

        response = self.client.post(
            reverse("unit_detail", kwargs={"pk": self.unit.pk}),
            {
                "label": "Unit 1A",
                "floor": "Floor 1",
                "area_sqm": "55.50",
                "unit_notes": "Unit note",
                "heating_type": "District heating",
                "boiler_installation_info": "Boiler info",
                "instant_water_heater_info": "No",
                "technical_notes": "Technical note",
                "rent_effective_start": "01.07.2026",
                "cold_rent": "1100.00",
                "utility_prepayment": "220.00",
                "total_rent": "",
                "rent_notes": "Rent note",
                "add_person_mode": "none",
                "existing_person": "",
                "first_name": "",
                "last_name": "",
                "email": "",
                "phone": "",
                "birthday": "",
                "relationship_notes": "",
                "new_role": LeasePerson.CO_TENANT,
                "new_move_in_date": "",
                "new_move_out_date": "",
                "new_notes": "",
                f"link_{link.pk}_first_name": "Ada",
                f"link_{link.pk}_last_name": "Lovelace",
                f"link_{link.pk}_email": "ada@example.com",
                f"link_{link.pk}_phone": "123",
                f"link_{link.pk}_birthday": "",
                f"link_{link.pk}_relationship_notes": "Primary contact",
                f"link_{link.pk}_role": LeasePerson.PRIMARY,
                f"link_{link.pk}_move_in_date": "01.01.2026",
                f"link_{link.pk}_move_out_date": "",
                f"link_{link.pk}_notes": "Signer",
            },
        )

        self.assertRedirects(response, reverse("unit_detail", kwargs={"pk": self.unit.pk}))
        self.unit.refresh_from_db()
        self.tenant.refresh_from_db()
        old_rent = self.lease.rent_periods.get(effective_start=date(2026, 1, 1))
        new_rent = self.lease.rent_periods.get(effective_start=date(2026, 7, 1))
        self.assertEqual(self.unit.label, "Unit 1A")
        self.assertEqual(self.unit.floor, "Floor 1")
        self.assertEqual(UnitTechnicalInfo.objects.get(unit=self.unit).heating_type, "District heating")
        self.assertEqual(self.tenant.email, "ada@example.com")
        self.assertEqual(old_rent.effective_end, date(2026, 6, 30))
        self.assertEqual(new_rent.total_rent, Decimal("1320.00"))

    def test_rent_period_delete_requires_confirmation_and_removes_only_selected_rent(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        old_rent = RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 1, 1), effective_end=date(2026, 6, 30), cold_rent=Decimal("1000"), utility_prepayment=Decimal("200"), total_rent=Decimal("1200"))
        keep_rent = RentPeriod.objects.create(lease=self.lease, effective_start=date(2026, 7, 1), cold_rent=Decimal("1100"), utility_prepayment=Decimal("220"), total_rent=Decimal("1320"))

        confirm = self.client.get(reverse("rent_period_delete", kwargs={"pk": old_rent.pk}))
        self.assertContains(confirm, "Delete Rent Period")
        self.assertContains(confirm, "permanently deletes")

        response = self.client.post(reverse("rent_period_delete", kwargs={"pk": old_rent.pk}))

        self.assertRedirects(response, f"{reverse('unit_detail', kwargs={'pk': self.unit.pk})}#history")
        self.assertFalse(RentPeriod.objects.filter(pk=old_rent.pk).exists())
        self.assertTrue(RentPeriod.objects.filter(pk=keep_rent.pk).exists())

    def test_loan_performance_exposes_nominal_and_effective_rates(self):
        loan = Loan.objects.create(property=self.property, name="Main loan", original_amount=Decimal("200000"))
        AnnualLoanSnapshot.objects.create(
            loan=loan,
            year=2026,
            opening_balance=Decimal("200000"),
            closing_balance=Decimal("194000"),
            interest_paid=Decimal("3600"),
            principal_paid=Decimal("6000"),
            interest_rate=Decimal("0.020000"),
        )

        row = loan_performance_rows(2026)[0]

        self.assertEqual(row.monthly_payment, Decimal("400.00"))
        self.assertEqual(row.effective_interest_rate, Decimal("0.0180"))
        self.assertEqual(row.amortization_rate, Decimal("0.0300"))
        self.assertEqual(row.interest_share_of_payment, Decimal("0.3750"))

    def test_current_debt_uses_closing_past_opening_future_and_approximation_current_year(self):
        self.assertEqual(
            current_debt_from_annual_snapshot(Decimal("100000.00"), Decimal("88000.00"), 2025, as_of=date(2026, 5, 25)),
            Decimal("88000.00"),
        )
        self.assertEqual(
            current_debt_from_annual_snapshot(Decimal("100000.00"), Decimal("88000.00"), 2027, as_of=date(2026, 5, 25)),
            Decimal("100000.00"),
        )
        self.assertEqual(
            current_debt_from_annual_snapshot(Decimal("100000.00"), Decimal("88000.00"), 2026, as_of=date(2026, 7, 2)),
            Decimal("94000.00"),
        )

    def test_loan_list_shows_all_kpis_and_current_debt_without_monthly_payment(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        loan = Loan.objects.create(property=self.property, name="Main loan", original_amount=Decimal("200000"))
        AnnualLoanSnapshot.objects.create(
            loan=loan,
            year=2025,
            opening_balance=Decimal("200000"),
            closing_balance=Decimal("194000"),
            interest_paid=Decimal("3600"),
            principal_paid=Decimal("6000"),
            debt_service=Decimal("9600"),
            monthly_payment=Decimal("800"),
            interest_rate=Decimal("0.020000"),
        )

        response = self.client.get(reverse("loan_list"), {"year": "2025"})
        edit_response = self.client.get(f"{reverse('loan_update', kwargs={'pk': loan.pk})}?year=2025")

        self.assertContains(response, "Your Current Debt")
        self.assertContains(response, "Showing your share")
        self.assertContains(response, "Effective Interest")
        self.assertContains(edit_response, "Showing full property values")
        self.assertContains(response, "Amortization Rate")
        self.assertNotContains(response, "Show more loan KPIs")
        self.assertNotContains(response, "Your Opening Debt")
        self.assertNotContains(response, "Your Closing Debt")
        self.assertNotContains(response, "Your Monthly Payment")
        self.assertNotContains(response, "Your Monthly Payments")
        self.assertNotContains(response, "€800.00")

    def test_monthly_payment_can_drive_annual_debt_service(self):
        loan = Loan.objects.create(property=self.property, name="Main loan", original_amount=Decimal("200000"))
        snapshot = AnnualLoanSnapshot.objects.create(
            loan=loan,
            year=2026,
            opening_balance=Decimal("200000"),
            closing_balance=Decimal("194000"),
            interest_paid=Decimal("3600"),
            principal_paid=Decimal("6000"),
            interest_rate=Decimal("0.020000"),
            monthly_payment=Decimal("900.00"),
            debt_service=Decimal("0.00"),
        )

        self.assertEqual(snapshot.debt_service, Decimal("10800.00"))

    def test_loan_balance_table_calculates_partial_year_values(self):
        form = LoanBalanceTableForm(
            data={
                "property": self.property.pk,
                "name": "Main loan",
                "original_amount": "100000.00",
                "start_date": "2026-03-15",
                "selected_year": "2026",
                "default_interest_rate": "0.020000",
                "default_monthly_payment": "1000.00",
                "closing_balance_2026": "92000.00",
                "lender": "",
                "maturity_date": "",
                "rate_reset_date": "",
                "notes": "",
            },
            selected_year=2026,
        )

        self.assertTrue(form.is_valid(), form.errors)
        loan, _ = form.save()
        snapshot = loan.annual_snapshots.get(year=2026)

        self.assertEqual(loan.original_amount, Decimal("100000.00"))
        self.assertEqual(loan.default_interest_rate, Decimal("0.020000"))
        self.assertEqual(loan.default_monthly_payment, Decimal("1000.00"))
        self.assertEqual(snapshot.opening_balance, Decimal("100000.00"))
        self.assertEqual(snapshot.debt_service, Decimal("10000.00"))
        self.assertEqual(snapshot.principal_paid, Decimal("8000.00"))
        self.assertEqual(snapshot.interest_paid, Decimal("2000.00"))
        self.assertEqual(form.rows[0]["active_months"], 10)

    def test_loan_balance_table_uses_prior_entered_closing_balance(self):
        form = LoanBalanceTableForm(
            data={
                "property": self.property.pk,
                "name": "Main loan",
                "original_amount": "100000.00",
                "start_date": "2026-01-01",
                "selected_year": "2027",
                "default_interest_rate": "0.000000",
                "default_monthly_payment": "1000.00",
                "closing_balance_2026": "90000.00",
                "closing_balance_2027": "80000.00",
                "lender": "",
                "maturity_date": "",
                "rate_reset_date": "",
                "notes": "",
            },
            selected_year=2027,
        )

        self.assertTrue(form.is_valid(), form.errors)
        loan, _ = form.save()
        snapshot = loan.annual_snapshots.get(year=2027)

        self.assertEqual(snapshot.opening_balance, Decimal("90000.00"))
        self.assertEqual(snapshot.principal_paid, Decimal("10000.00"))
        self.assertEqual(snapshot.interest_paid, Decimal("2000.00"))

    def test_lower_than_expected_closing_balance_becomes_extra_repayment(self):
        form = LoanBalanceTableForm(
            data={
                "property": self.property.pk,
                "name": "Main loan",
                "original_amount": "100000.00",
                "start_date": "2026-01-01",
                "selected_year": "2026",
                "default_interest_rate": "0.000000",
                "default_monthly_payment": "1000.00",
                "closing_balance_2026": "85000.00",
                "lender": "",
                "maturity_date": "",
                "rate_reset_date": "",
                "notes": "",
            },
            selected_year=2026,
        )

        self.assertTrue(form.is_valid(), form.errors)
        row = form.rows[0]
        self.assertEqual(row["expected_closing_balance"], Decimal("88000.00"))
        self.assertEqual(row["extra_repayment"], Decimal("3000.00"))
        self.assertEqual(row["debt_service"], Decimal("15000.00"))
        self.assertEqual(row["principal_paid"], Decimal("15000.00"))
        self.assertEqual(row["interest_paid"], Decimal("0.00"))

    def test_higher_than_expected_closing_balance_shows_variance(self):
        form = LoanBalanceTableForm(
            data={
                "property": self.property.pk,
                "name": "Main loan",
                "original_amount": "100000.00",
                "start_date": "2026-01-01",
                "selected_year": "2026",
                "default_interest_rate": "0.000000",
                "default_monthly_payment": "1000.00",
                "closing_balance_2026": "90000.00",
                "lender": "",
                "maturity_date": "",
                "rate_reset_date": "",
                "notes": "",
            },
            selected_year=2026,
        )

        self.assertTrue(form.is_valid(), form.errors)
        row = form.rows[0]
        self.assertEqual(row["expected_closing_balance"], Decimal("88000.00"))
        self.assertEqual(row["variance"], Decimal("2000.00"))
        self.assertEqual(row["extra_repayment"], Decimal("0.00"))
        self.assertEqual(row["debt_service"], Decimal("12000.00"))
        self.assertEqual(row["principal_paid"], Decimal("10000.00"))
        self.assertEqual(row["interest_paid"], Decimal("2000.00"))

    def test_loan_balance_table_rejects_closing_above_opening(self):
        form = LoanBalanceTableForm(
            data={
                "property": self.property.pk,
                "name": "Main loan",
                "original_amount": "100000.00",
                "start_date": "2026-01-01",
                "selected_year": "2026",
                "default_interest_rate": "0.000000",
                "default_monthly_payment": "100.00",
                "closing_balance_2026": "101000.00",
                "lender": "",
                "maturity_date": "",
                "rate_reset_date": "",
                "notes": "",
            },
            selected_year=2026,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("cannot exceed opening", str(form.errors))

    def test_loan_balance_table_edit_view_updates_loan_and_year_data(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        loan = Loan.objects.create(property=self.property, name="Old name", original_amount=Decimal("100000.00"), start_date=date(2026, 1, 1), default_interest_rate=Decimal("0.000000"), default_monthly_payment=Decimal("1000.00"))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2026, opening_balance=Decimal("100000.00"), closing_balance=Decimal("95000.00"), monthly_payment=Decimal("1000.00"), debt_service=Decimal("12000.00"), principal_paid=Decimal("5000.00"), interest_paid=Decimal("7000.00"))

        response = self.client.post(
            f"{reverse('loan_update', kwargs={'pk': loan.pk})}?year=2026",
            {
                "property": self.property.pk,
                "name": "Updated loan",
                "original_amount": "100000.00",
                "start_date": "2026-01-01",
                "selected_year": "2026",
                "default_interest_rate": "0.000000",
                "default_monthly_payment": "1100.00",
                "closing_balance_2026": "92000.00",
                "lender": "Bank",
                "maturity_date": "",
                "rate_reset_date": "",
                "notes": "Updated note",
            },
        )

        self.assertRedirects(response, f"{reverse('loan_list')}?year=2026")
        loan.refresh_from_db()
        snapshot = loan.annual_snapshots.get(year=2026)
        self.assertEqual(loan.name, "Updated loan")
        self.assertEqual(loan.lender, "Bank")
        self.assertEqual(loan.default_monthly_payment, Decimal("1100.00"))
        self.assertEqual(snapshot.opening_balance, Decimal("100000.00"))
        self.assertEqual(snapshot.debt_service, Decimal("13200.00"))
        self.assertEqual(snapshot.principal_paid, Decimal("8000.00"))
        self.assertEqual(snapshot.interest_paid, Decimal("5200.00"))

    def test_loan_list_links_loan_name_to_unified_edit_page(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        loan = Loan.objects.create(property=self.property, name="Main loan", original_amount=Decimal("100000.00"), start_date=date(2026, 1, 1))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2026, opening_balance=Decimal("100000.00"), closing_balance=Decimal("95000.00"), monthly_payment=Decimal("1000.00"), debt_service=Decimal("12000.00"), principal_paid=Decimal("5000.00"), interest_paid=Decimal("7000.00"))

        response = self.client.get(f"{reverse('loan_list')}?year=2026")

        self.assertContains(response, f'href="{reverse("loan_update", kwargs={"pk": loan.pk})}?year=2026"')
        self.assertContains(response, "Main loan")

    def test_loan_edit_accepts_german_start_date_and_creates_missing_years(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        loan = Loan.objects.create(property=self.property, name="Heestweg loan", original_amount=Decimal("100000.00"), start_date=date(2026, 1, 1), default_interest_rate=Decimal("0.000000"), default_monthly_payment=Decimal("1000.00"))
        AnnualLoanSnapshot.objects.create(loan=loan, year=2026, opening_balance=Decimal("100000.00"), closing_balance=Decimal("95000.00"), monthly_payment=Decimal("1000.00"), debt_service=Decimal("12000.00"), principal_paid=Decimal("5000.00"), interest_paid=Decimal("7000.00"))

        response = self.client.post(
            f"{reverse('loan_update', kwargs={'pk': loan.pk})}?year=2026",
            {
                "property": self.property.pk,
                "name": "Heestweg loan",
                "original_amount": "100000.00",
                "start_date": "30.11.2021",
                "selected_year": "2026",
                "default_interest_rate": "0.000000",
                "default_monthly_payment": "1000.00",
                "closing_balance_2026": "95000.00",
                "lender": "",
                "maturity_date": "",
                "rate_reset_date": "",
                "notes": "",
            },
        )

        self.assertRedirects(response, f"{reverse('loan_list')}?year=2026")
        loan.refresh_from_db()
        self.assertEqual(loan.start_date, date(2021, 11, 30))
        self.assertEqual(list(loan.annual_snapshots.order_by("year").values_list("year", flat=True)), [2021, 2022, 2023, 2024, 2025, 2026])
        self.assertEqual(loan.annual_snapshots.get(year=2021).opening_balance, Decimal("100000.00"))
        self.assertEqual(loan.annual_snapshots.get(year=2021).debt_service, Decimal("2000.00"))

    def test_failed_loan_save_shows_feedback(self):
        user = User.objects.create_user(username="admin", password="secret")
        self.client.force_login(user)
        loan = Loan.objects.create(property=self.property, name="Main loan", original_amount=Decimal("100000.00"), start_date=date(2026, 1, 1), default_interest_rate=Decimal("0.000000"), default_monthly_payment=Decimal("1000.00"))

        response = self.client.post(
            f"{reverse('loan_update', kwargs={'pk': loan.pk})}?year=2026",
            {
                "property": self.property.pk,
                "name": "Main loan",
                "original_amount": "100000.00",
                "start_date": "2026-01-01",
                "selected_year": "2026",
                "default_interest_rate": "0.000000",
                "default_monthly_payment": "1000.00",
                "closing_balance_2026": "101000.00",
                "lender": "",
                "maturity_date": "",
                "rate_reset_date": "",
                "notes": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Save failed. Please check the highlighted fields.")
        self.assertContains(response, "Closing balance cannot exceed opening balance.")
